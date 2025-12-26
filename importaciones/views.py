import hashlib
import math
import io
import requests
import random
import zipfile
from collections import OrderedDict
import traceback
from pathlib import Path
import openpyxl
import fitz
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import connections,IntegrityError,transaction
from django.db.models import Q, Count, Min
from django.utils.encoding import smart_str
from django.utils.timezone import make_aware
from pypdf import PdfReader, PdfWriter
from reportlab.lib.styles import getSampleStyleSheet
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from reportlab.lib.units import cm
from rest_framework import generics, status, permissions, viewsets
from rest_framework.permissions import IsAuthenticated, AllowAny

from .models import (OrdenCompraStarsoft, GastosExtra, Proveedor, OrdenCompraDespacho, Empresa, OrdenCompra, Producto,
                     ProveedorTransporte, Transportista,
                     Despacho, DetalleDespacho, ConfiguracionDespacho, Declaracion, Documento, ExpedienteDeclaracion,
                     TipoDocumento)
from .forms import BaseDatosForm
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse, Http404, FileResponse
from rest_framework.decorators import api_view, permission_classes, action
import pandas as pd
import pytesseract
from PIL import Image
import pdfplumber
import json
from django.shortcuts import get_object_or_404
from .serializers import *
from .utils import renderizar_template, convertir_html_a_pdf, procesar_data_reporte, procesar_data_bd_reporte, \
    calcular_monto_descuento_estiba, calcular_peso_no_considerado_por_sacos_faltante, \
    calcular_diferencia_de_peso_por_cobrar_kg, calcular_costo_por_kg, calcular_descuento_sacos, \
    calcular_descuento_solo_sacos, calcular_monto_luego_dsctos_sacos,calcular_hash_archivo
import os
from reportlab.lib.pagesizes import landscape, A4, letter
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.colors import Color
import rarfile
#rarfile.UNRAR_TOOL = r"C:\Program Files\WinRAR\UnRAR.exe"
from django.conf import settings
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView
from django.contrib.contenttypes.models import ContentType
from .utilities.dto_despacho import *
import mimetypes
from usuarios.permissions import CanAccessImportaciones, IsImportacionesAdmin, CanEditDocuments, CanDeleteResource



def get_db_connection(base_datos):
    connection = connections[base_datos]  # Usa la base de datos dinámica
    return connection

def extract_pdf_tables(file):
    with pdfplumber.open(file) as pdf:
        tables = []

        # Intentar extraer tablas de cada página del PDF
        for i, page in enumerate(pdf.pages):
            table = page.extract_table()

            if table:
                print(f"Tabla extraída en página {i + 1}:")
                headers = table[0]  # La primera fila contiene los encabezados
                table_data = [headers]  # Agregar los encabezados como la primera fila

                for row in table[1:]:
                    table_data.append(row)  # Agregar las filas de datos

                tables.append(table_data)
            else:
                print(f"No se encontró una tabla en la página {i + 1}")

        # Si no se extrajeron tablas, intentamos extraer texto y convertirlo
        if not tables:
            print("No se encontraron tablas, pero aquí está el texto extraído:")
            text = ""
            for page in pdf.pages:
                text += page.extract_text()  # Extraer todo el texto del PDF

            #print(text)
            lines = text.splitlines()  # Dividir el texto en líneas

            data = []
            header_found = False

            # Encabezado ajustado según la estructura del texto
            header = ["No.","NRO CE", "FECHA", "PLACA", "TRANSPORTISTA", "BULTOS", "PESO BRUTO", "PESO NETO",
                      "Nros. Precintos SENASAG"]

            for line in lines:
                # Detectar si la línea es parte de la tabla (empieza con el encabezado)
                if line.startswith("No.") and not header_found:
                    header_found = True  # Marca que el encabezado ha sido encontrado
                elif header_found:
                    # Si encontramos la línea "TOTALES", detenemos la extracción sin agregar la línea
                    if "TOTALES" in line:
                        break
                    else:
                        # Separar las columnas por espacios
                        columns = line.split()

                        # Asegurarnos de tener al menos las 8 columnas esperadas antes de procesar
                        if len(columns) >= len(header):
                            # Organizar las columnas fijas
                            no = columns[0]  # No.
                            nro_ce = columns[1]  # NRO CE
                            fecha = columns[2]  # FECHA
                            placa = columns[3]  # PLACA

                            # Las últimas 4 columnas son siempre los mismos datos
                            bultos = columns[-4]
                            peso_bruto = columns[-3]
                            peso_neto = columns[-2]
                            nro_precintos = columns[-1]

                            # Todo lo que queda en el medio se considera como el nombre del transportista
                            transportista = " ".join(columns[4:-4])

                            # Crear la fila de datos organizada
                            row = [no, nro_ce, fecha, placa, transportista, bultos, peso_bruto, peso_neto,
                                   nro_precintos]
                            data.append(row)
            #print("encontro encabezado?", header_found)
            # Crear un DataFrame con los datos extraídos
            df_text_table = pd.DataFrame(data, columns=header)
            # Convertir el DataFrame a una lista de diccionarios
            return df_text_table.to_dict(orient='records')

def extract_excel_data(file):
    try:
        df = pd.read_excel(file, sheet_name=None)
        data = {}
        for sheet_name, sheet_data in df.items():
            data[sheet_name] = sheet_data.to_dict(orient='records')  # Convertir a lista de diccionarios
        return data
    except Exception as e:
        return {"error": str(e)}

def extract_image_text(file):
    img = Image.open(file)
    text = pytesseract.image_to_string(img)
    return text

@api_view(['POST'])
def upload_file_excel(request):
    try:
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({"error": "No se proporcionó archivo."}, status=400)

        # Leer el archivo Excel usando openpyxl
        df = pd.read_excel(uploaded_file, engine='openpyxl', skiprows=1, header=0)

        # Limpiar y renombrar columnas
        df.columns = ["Item", "PLACA", "sacos_cargados", "peso_cargado", "PLACA_Balanza",
                      "sacos_entregados", "peso_balanza", "Merma", "descuentos_faltantes",
                      "descuentos_rotos", "descuentos_humedos", "descuentos_mojados", "Status_pago"]

        # Renombrar columnas para el frontend
        df.rename(columns={
            "Item": "item_nombre",
            "PLACA": "placa_salida",
            "sacos_cargados": "sacos_cargados",
            "peso_cargado": "peso_salida",
            "PLACA_Balanza": "placa_llegada",
            "sacos_entregados": "sacos_descargados",
            "peso_balanza": "peso_llegada",
            "Merma": "merma",
            "descuentos_faltantes": "sacos_faltantes",
            "descuentos_rotos": "sacos_rotos",
            "descuentos_humedos": "sacos_humedos",
            "descuentos_mojados": "sacos_mojados",
            "Status_pago": "descripcion_pago"
        }, inplace=True)



        # Procesar códigos usando la leyenda
        def procesar_codigo(codigo):
            try:
                codigo_entero = int(math.floor(float(codigo)))
                return leyenda_codigos.get(codigo_entero, "Seleccione")
            except (ValueError, TypeError):
                return "Seleccione"

        # Aplicar procesamiento de códigos
        df["pago_estiba"] = df["descripcion_pago"].apply(procesar_codigo)

        # Borrar columnas que no necesito
        df.drop(columns=["item_nombre", "merma","descripcion_pago"], inplace=True)

        # Convertir el DataFrame en una respuesta JSON
        result = df.to_dict(orient='records')

        return JsonResponse({"status":"success","tabla": result}, status=200)

    except Exception as e:
        return JsonResponse({"status":"error","error": f"Error al procesar archivo: {str(e)}"}, status=400)

@api_view(['POST'])
def upload_file(request):
    if 'file' in request.FILES:
        file = request.FILES['file']
        file_type = file.name.split('.')[-1].lower()

        try:
            if file_type == 'pdf':
                tables = extract_pdf_tables(file)
                response_data = {"tables": tables}
            elif file_type in ['xls', 'xlsx']:
                data = extract_excel_data(file)
                response_data = {"data": data}
            elif file_type in ['jpg', 'jpeg', 'png']:
                text = extract_image_text(file)
                response_data = {"text": text}
            else:
                response_data = {'message': 'Formato de archivo no soportado'}

            return JsonResponse({'status': 'success', 'data': response_data}, status=200)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({"error": "No se proporcionó archivo."}, status=400)

leyenda_codigos = {
    1: "Transbordo",
    2: "Pago estiba",
    3: "No pago estiba",
    4: "Pago parcial",
}

@csrf_exempt
def listar_importaciones(request):
    form = BaseDatosForm(request.POST or None)
    registros = None

    if request.method == 'POST' and form.is_valid():
        base_datos = form.cleaned_data['base_datos']
        registros = OrdenCompraStarsoft.objects.using(base_datos).all()  # Consultar en la base seleccionada

    return render(request, 'importaciones/lista_importaciones.html', {'form': form, 'registros': registros})

@csrf_exempt
def buscar_orden_importacion(request):
    base_datos = request.GET.get('base_datos', 'default')  # Obtén el parámetro base_datos desde la URL
    query = request.GET.get('query', '')  # El término de búsqueda (CNUMERO)

    # Verifica que los parámetros sean válidos
    if base_datos and query:
        try:
            # Obtener la conexión a la base de datos
            connection = get_db_connection(base_datos)

            # Realizar la consulta a la base de datos
            with connection.cursor() as cursor:
                cursor.execute(f"""
                    SELECT TOP 5 
                        i.CNUMERO,
                        i.CDESARTIC,
                        i.CCODARTIC,
                        i.NCANTIDAD,
                        i.CUNIDAD,
                        i.NPREUNITA,
                        i.NTOTVENT,
                        o.CDESPROVE,
                        o.CCODPROVE
                    FROM IMPORD i
                    INNER JOIN IMPORC o ON i.CNUMERO = o.CNUMERO
                    WHERE i.CNUMERO LIKE %s
                """, [f'%{query}%'])
                resultados = cursor.fetchall()

            # Serializar los resultados
            resultado_serializado = []

            for row in resultados:

                orden = {
                    'numero_oc': row[0],  # Asegúrate de que el índice coincida con el esquema de tu tabla
                    'producto': row[1],  # Este es solo un ejemplo, ajusta según las columnas
                    'codigo_producto': row[2],
                    'cantidad': row[3],
                    'unidad_medida': row[4],
                    'precio_unitario': row[5],
                    'precio_total': row[6],
                    'proveedor': row[7],  # Ajusta según la columna correcta
                    'codprovee': row[8],  # Ajusta según la columna correcta
                }
                resultado_serializado.append(orden)

            # Retornar los resultados en formato JSON
            return JsonResponse(resultado_serializado, safe=False, status=200)

        except ValueError as e:
            # Si hay un error con la base de datos
            return JsonResponse({'error': str(e)}, status=500)

        except Exception as e:
            # Manejo de errores generales
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Parámetros inválidos'}, status=400)

@csrf_exempt
def buscar_proveedor(request):
    if request.method == 'GET':
        base_datos = request.GET.get('base_datos')  # Base de datos seleccionada
        termino_busqueda = request.GET.get('query')  # Término de búsqueda

        # Valida que se haya proporcionado la base de datos y el término de búsqueda
        if base_datos and termino_busqueda:
            # Filtra las órdenes que coincidan con el término de búsqueda
            registros = Proveedor.objects.using(base_datos).filter(
                PRVCNOMBRE__icontains=termino_busqueda
            )[:5]  # Carga los detalles relacionados

            # Serializa las órdenes junto con sus detalles
            resultado = []
            for registro in registros:
                # Extraemos la orden principal
                proveedor = {
                    'nombre': registro.PRVCNOMBRE,
                }

                #print(proveedor)
                resultado.append(proveedor)

            return JsonResponse(resultado, safe=False, status=200)
        else:
            return JsonResponse({'error': 'Parámetros inválidos'}, status=400)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

def generar_reporte_base(request):
    data_unprocess = json.loads(request.body)
    data = procesar_data_reporte(data_unprocess)
    buffer = io.BytesIO()

    # Crear un archivo PDF
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    c.setTitle(" Reporte cálculo de flete detallado")

    # Obtener las dimensiones de la página
    width, height = landscape(A4)

    # primera linea
    c.setFont("Helvetica", 10)
    current_y = height - 40
    c.drawString(40, current_y, f"{data['procesado']['empresa']}")

    current_datetime = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    c.setFont("Helvetica", 7)
    c.drawString(width - 130, current_y, f"{current_datetime}")

    # segunda linea
    c.setFont("Helvetica-Bold", 10)
    texto = f"{data['dataForm']['producto']}"
    text_width = c.stringWidth(texto, "Helvetica", 12)
    # Calcular la posición X para centrar el texto
    x_position = (width - text_width) / 2 + 20
    # Dibujar el texto centrado
    current_y -= 20
    c.drawString(x_position, current_y, texto)

    # tercera linea
    current_y -= 20
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, current_y, f"CARTA PORTE: {data['dataForm']['cartaPorte']}")

    # CUARTA LINEA DATOS DEL DESPACHO
    ordenes_recojo = data['dataForm']['ordenRecojo']
    ordenes_string = " / ".join([f"{row['oc']['numero_oc']}({row['numeroRecojo']})" for row in ordenes_recojo])
    textos = [
        f"N° DUA: {data['dataForm']['dua']}",
        f"Fec. Num.: {data['procesado']['fecha_numeracion']}",
        f"Factura N°: {data['dataForm']['numFactura']}",
        f"OC: {ordenes_string}",
        f"CANT. {data['procesado']['total_sacos_cargados']} sacos",
        f" {data['dataForm']['pesoNetoCrt']} kg"
    ]
    # Fuente y tamaño
    font_name = "Helvetica"
    font_size = 8
    c.setFont(font_name, font_size)
    # Espacio entre columnas
    margin = 30
    # Posición inicial en Y (mantener fija)
    current_y -= 20
    y_position = current_y
    # Posición inicial en X
    x_position = 40
    # Dibujar todos los textos en una línea tomando en cuenta su tamaño para darle cierto espaciado
    for texto in textos:
        # Calcular el ancho de cada texto
        text_width = c.stringWidth(texto, font_name, font_size)

        # Dibujar el texto
        c.drawString(x_position, y_position, texto)

        # Actualizar la posición X para el siguiente texto
        x_position += text_width + margin

        # Si el texto excede el tamaño de la página en el eje X, detener la creación del contenido
        if x_position > width:
            break

    # QUINTA LINEA DONDE INICIA LA TABLA:
    current_y -= 20
    # Datos iniciales de la tabla #1 (Titulos y subtitulos)
    data_table = [
        # Línea de títulos
        ['FEC. INGRE.', 'EMPRESA DE TRANSP.', 'CARGA', '', '', '', 'DESCARGA', '', '', '', 'DESCUENTOS', '', '', '',
         ''],
        ['', '', 'N°', 'Placa S.', 'Sacos C.', 'Peso S.', 'Placa L.', 'Sacos D.', 'Peso L.', 'Merma', 'S. Falt.',
         'S. Rotos', 'S. Humed.', 'S. Mojad.', 'Estibaje'],
    ]
    # Agregamos datos dinamicamente a la tabla
    styles = getSampleStyleSheet()
    parrafo = Paragraph(f"<para align=center spaceb=3><b>{data['dataForm']['transportista']}</b></para>",
                        styles["BodyText"])
    for i, row in enumerate(data['dataTable']):
        if i == 0:  # Primera fila (datos especiales)
            data_table.append([
                f"{data['procesado']['fecha_numeracion']}",
                parrafo,
                str(row['numero']),
                row['placa'],
                str(row['sacosCargados']),
                str(row['pesoSalida']),
                row['placaLlegada'],
                str(row['sacosDescargados']),
                str(row['pesoLlegada']),
                str(row['merma']),
                str(row['sacosFaltantes']),
                str(row['sacosRotos']),
                str(row['sacosHumedos']),
                str(row['sacosMojados']),
                str(row['pagoEstiba']),
            ])
        else:  # Para el resto de las filas
            data_table.append([
                "",  # Cadena vacía
                "",  # Cadena vacía
                str(row['numero']),
                row['placa'],
                str(row['sacosCargados']),
                str(row['pesoSalida']),
                row['placaLlegada'],
                str(row['sacosDescargados']),
                str(row['pesoLlegada']),
                str(row['merma']),
                str(row['sacosFaltantes']),
                str(row['sacosRotos']),
                str(row['sacosHumedos']),
                str(row['sacosMojados']),
                str(row['pagoEstiba']),
            ])
    # Agregamos
    data_table.append(
        ['', f"Flete x TM: $ {data['dataForm']['fletePactado']:.2f}", f"{data['procesado']['len_tabla']}", 'TOTAL',
         f"{data['procesado']['total_sacos_cargados']}", f"{data['procesado']['suma_peso_salida']}", 'TOTAL',
         f"{data['procesado']['total_sacos_descargados']}", f"{data['procesado']['suma_peso_llegada']}",
         f"{data['procesado']['merma_total']}", f"{data['procesado']['total_sacos_faltantes']}",
         f"{data['procesado']['total_sacos_rotos']}", f"{data['procesado']['total_sacos_humedos']}",
         f"{data['procesado']['total_sacos_mojados']}"])
    # Ancho de columnas
    col_widths = [60, 110, 20, 60, 40, 60, 60, 60, 40, 60, 30, 30, 30, 30, 70]
    # Crear la tabla
    table = Table(data_table, colWidths=col_widths, rowHeights=15)
    color_rosa = Color(red=250 / 255, green=210 / 255, blue=202 / 255)
    color_celeste = Color(red=176 / 255, green=225 / 255, blue=250 / 255)
    numero_filas = len(data_table)
    # Estilos para la tabla
    table_style = TableStyle([
        ('SPAN', (0, 0), (0, 1)),  # Fusionar "FEC. INGRE."
        ('SPAN', (1, 0), (1, 1)),  # Fusionar "EMPRESA DE TRANSP."
        ('SPAN', (2, 0), (5, 0)),  # Fusionar "CARGA"
        ('SPAN', (6, 0), (9, 0)),  # Fusionar "DESCARGA"
        ('SPAN', (10, 0), (14, 0)),  # Fusionar "DESCUENTOS"
        ('SPAN', (0, 2), (0, numero_filas - 1)),  # Fusionar "fecha"
        ('SPAN', (1, 2), (1, numero_filas - 2)),  # Fusionar "nombre empresa transporte"
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Centrar todo
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Fondo gris para los títulos
        ('BACKGROUND', (0, 1), (-1, 1), colors.lightgrey),  # Fondo gris claro para subtítulos
        ('BACKGROUND', (2, 0), (5, numero_filas), colors.yellowgreen),  # Color "CARGA"
        ('BACKGROUND', (6, 0), (9, numero_filas), color_celeste),  # Color "DESCARGA"
        ('BACKGROUND', (10, 0), (14, numero_filas), color_rosa),  # Fusionar "DESCUENTOS"
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Negrita en títulos
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),  # Negrita en subtítulos

        # ('FONTSIZE', (15, 0), (15, numero_filas - 1), 5),
    ])
    table_style.add('FONTSIZE', (1, 2), (1, numero_filas - 2), 5)
    # Aplicar estilos a la tabla
    table.setStyle(table_style)
    # Calcular la altura total de la tabla
    table_height = len(data_table) * 15  # 20 es la altura de cada fila
    # Calcular la posición de la tabla en la página
    table.wrapOn(c, width, height)
    table.drawOn(c, 40, current_y - table_height)  # Ajustar la posición dependiendo de la cantidad de filas

    current_y = current_y - table_height - 20

    # Data para la tabla #2 (Detalles de pesos)
    second_data_table = [
        ["Faltante peso Sta. Cruz - Desaguadero:", f"{data['procesado']['diferencia_de_peso']:.2f} Kg."],
        ["Merma permitida:", f"{data['dataExtra']['mermaPermitida']:.2f} Kg."],
        ["Desc. diferencia de peso:", f"{data['procesado']['diferencia_peso_por_cobrar']:.2f} Kg."],
        ["Desc. sacos falantes:", f"{data['procesado']['descuento_peso_sacos_faltantes']:.2f} Kg."]
    ]
    second_col_widths = [110, 50]
    second_table = Table(second_data_table, colWidths=second_col_widths, rowHeights=14)
    second_table_style = TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Centrar
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
    ])
    second_table.setStyle(second_table_style)
    second_table_height = len(second_data_table) * 14
    second_table.wrapOn(c, width, height)
    second_table.drawOn(c, 40, current_y - second_table_height)

    # Data para la tabla #3 (Resumen de descuentos)
    third_data_table = [
        ["FLETE", f'$ {data["procesado"]["flete_base"]:.2f}'],
        ["Dscto. por dif. de peso", f'$ {data["procesado"]["descuento_por_diferencia_peso"]:.2f}'],
        ["Dscto. por sacos faltantes", f'$ {data["procesado"]["descuento_sacos_faltantes"]:.2f}'],
        ["Dscto. por sacos rotos", f'$ {data["procesado"]["descuento_sacos_rotos"]:.2f}'],
        ["Dscto. por sacos humedos", f'$ {data["procesado"]["descuento_sacos_humedos"]:.2f}'],
        ["Dscto. por sacos mojados", f'$ {data["procesado"]["descuento_sacos_mojados"]:.2f}'],
        ["FLETE TOTAL", f'$ {data["procesado"]["total_luego_dsctos_sacos"]:.2f}'],
    ]
    # posicion en x para la tercera tabla:
    x_position_for_third_table = 260
    third_col_widths = [110, 40]
    third_table = Table(third_data_table, colWidths=third_col_widths, rowHeights=14)
    third_table_style = TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
    ])
    third_table.setStyle(third_table_style)
    third_table_height = len(third_data_table) * 14
    third_table.wrapOn(c, width, height)
    third_table.drawOn(c, x_position_for_third_table, current_y - third_table_height)

    # PARA VALIDACION DEL VALOR CRT
    peso_total_toneladas = data['dataForm']['pesoNetoCrt'] / 1000
    flete_pactado = data["dataForm"]["fletePactado"]
    monto_flete = round((peso_total_toneladas * flete_pactado), 2)
    flete_base = data["procesado"]["flete_base"]
    estado_crt = ""
    if (flete_base <= monto_flete):
        estado_crt = "NO SOBREPASA VALOR CRT"
    else:
        estado_crt = "SOBREPASA VALOR CRT"
    c.setFont("Helvetica-Bold", 7)
    c.drawString(x_position_for_third_table + 150 + 10, current_y - 8, estado_crt)

    # Primera LLave para descuento sacos RMH
    c.setLineWidth(0.5)
    first_x_position_for_key = x_position_for_third_table + 150 + 10
    first_key_y_position = current_y - (14 * 3)
    last_key_y_position = first_key_y_position - (14 * 3)
    middle_of_key = (first_key_y_position + last_key_y_position) / 2
    inclinacion = 2
    c.line(first_x_position_for_key, first_key_y_position - inclinacion, first_x_position_for_key,
           last_key_y_position + inclinacion)
    c.line(first_x_position_for_key, first_key_y_position - inclinacion, first_x_position_for_key - inclinacion,
           first_key_y_position)
    c.line(first_x_position_for_key - inclinacion, last_key_y_position, first_x_position_for_key,
           last_key_y_position + inclinacion)
    c.line(first_x_position_for_key, middle_of_key, first_x_position_for_key + inclinacion, middle_of_key)
    c.setFont("Helvetica-Bold", 7)
    c.drawString(first_x_position_for_key + inclinacion + 4, middle_of_key,
                 f"$ {data['procesado']['total_descuento_solo_sacos']:.2f}")

    # Segunda llave para descuento de sacos RMH
    c.setLineWidth(0.5)

    # Coordenadas para dibujar segunda la llave
    key_start_x = first_x_position_for_key + inclinacion + 4 + 40
    key_start_y = current_y - 14
    key_end_y = key_start_y - (14 * 5)
    key_middle_y = (key_start_y + key_end_y) / 2
    key_inclination = 2

    # Dibujar la llave
    c.line(key_start_x, key_start_y - key_inclination, key_start_x, key_end_y + key_inclination)
    c.line(key_start_x, key_start_y - key_inclination, key_start_x - key_inclination, key_start_y)
    c.line(key_start_x - key_inclination, key_end_y, key_start_x, key_end_y + key_inclination)
    c.line(key_start_x, key_middle_y, key_start_x + key_inclination, key_middle_y)

    # Texto del descuento
    c.setFont("Helvetica-Bold", 7)
    c.drawString(key_start_x + key_inclination + 4, key_middle_y,
                 f"$ {data['procesado']['aux_descuento']:.2f}")

    # Lista de placas y estado de estiba:
    y_position_for_list = current_y - third_table_height - 10
    c.setFont("Helvetica", 6)
    for item in data["procesado"]["pago_estiba_list"]:
        # Crear texto en una línea
        c.drawString(x_position_for_third_table, y_position_for_list,
                     f"{item['placa']} {item['detalle']}")  # Placa y detalle
        c.drawString(x_position_for_third_table + 130, y_position_for_list,
                     f"$ {item['monto_descuento']:.2f}")  # Monto alineado a la derecha
        y_position_for_list -= 10  # Reducir la posición vertical para la próxima línea

    # Coordenadas para dibujar la tercera llave
    tirth_key_start_x = first_x_position_for_key
    tirth_key_start_y = current_y - third_table_height - 5
    tirth_key_end_y = y_position_for_list + 10
    tirth_key_middle_y = (tirth_key_start_y + tirth_key_end_y) / 2
    tirth_key_inclination = 2
    # Dibujar la tercera llave
    c.line(tirth_key_start_x, tirth_key_start_y - tirth_key_inclination, tirth_key_start_x,
           tirth_key_end_y + tirth_key_inclination)
    c.line(tirth_key_start_x, tirth_key_start_y - tirth_key_inclination, tirth_key_start_x - tirth_key_inclination,
           tirth_key_start_y)
    c.line(tirth_key_start_x - tirth_key_inclination, tirth_key_end_y, tirth_key_start_x,
           tirth_key_end_y + tirth_key_inclination)
    c.line(tirth_key_start_x, tirth_key_middle_y, tirth_key_start_x + tirth_key_inclination, tirth_key_middle_y)
    # Texto del descuento
    c.setFont("Helvetica-Bold", 7)
    c.drawString(tirth_key_start_x + tirth_key_inclination + 4, tirth_key_middle_y - 2,
                 f"$ {data['procesado']['total_descuento_estiba']:.2f}")

    # Escribir linea de neto a pagar
    table_total_data = [
        ["TOTAL A PAGAR:", f"${data['procesado']['total_a_pagar']:.2f}"]
    ]
    # Crear la tabla
    table_total = Table(table_total_data, colWidths=[110, 40])  # Anchos de columnas
    # Aplicar estilo a la tabla
    table_total.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.orange),  # Fondo naranja
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # Texto en color negro
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alineación izquierda primera columna
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),  # Fuente en negrita
        ('FONTSIZE', (0, 0), (-1, -1), 8),  # Tamaño de la fuente
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alineación derecha segunda columna
    ]))
    altura_tabla_total = len(table_total_data) * 14
    table_total.wrapOn(c, width, height)
    table_total.drawOn(c, x_position_for_third_table, y_position_for_list - altura_tabla_total - 20)



    # Data para la tabla #4 (resumen de descuentos)
    new_x_position = 260 + 150 + 150
    fourth_data_table = [
        ["FLETE", f"$ {data['procesado']['flete_base']:.2f}"],
        ["Dscto.", f"$ {data['procesado']['total_dsct']:.2f}"],
        ["Neto.", f"$ {data['procesado']['total_a_pagar']:.2f}"],
    ]
    fourth_col_widths = [30, 50]
    fourth_table = Table(fourth_data_table, colWidths=fourth_col_widths, rowHeights=14)
    fourth_table_style = TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
    ])
    fourth_table.setStyle(fourth_table_style)
    fourth_table_height = len(fourth_data_table) * 14
    fourth_table.wrapOn(c, width, height)
    fourth_table.drawOn(c, new_x_position, current_y - fourth_table_height - (14 * 3))
    c.setFont("Helvetica-Bold", 7)
    c.drawString(new_x_position, current_y - fourth_table_height + 5, "RESUMEN")

    # Data para la tabla #5 (determinacion de costos)
    second_new_x_position = new_x_position + 120
    c.setFont("Helvetica-Bold", 6)
    c.drawString(second_new_x_position, current_y + 5, "Determinacion de costo x Kg.")
    fifth_data_table = [
        ["C.P", f"{data['procesado']['precio_por_tonelada']:.2f}"],
        ["F.B", f"$ {data['dataForm']['fletePactado']:.2f}"],
        ["G.N", f"{data['dataExtra']['gastosNacionalizacion']}"],
        ["MF CIA", f"{data['dataExtra']['margenFinanciero']}"],
        ["IGV 18%", f"{data['procesado']['igv']}"],
        ["PRECIO DES", f"{data['procesado']['precio_bruto_final']}"],
        ["COSTO TM", f"{data['procesado']['precio_por_tonelada_final']}"],
        ["COSTO Kg", f"{data['procesado']['precio_por_kg_final']}"],
    ]
    fifth_col_widths = [50, 40]
    fifth_table = Table(fifth_data_table, colWidths=fifth_col_widths, rowHeights=14)
    fifth_table_style = TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
    ])
    fifth_table.setStyle(fifth_table_style)
    fifth_table_height = len(fifth_data_table) * 14
    fifth_table.wrapOn(c, width, height)
    fifth_table.drawOn(c, second_new_x_position, current_y - fifth_table_height)

    # SALTO DE LINEA PARA ESPACIO DE DESCUENTOS
    current_y = current_y - second_table_height - 20
    x_start = 40  # Eje X constante
    line_height = 10  # Interlineado
    # Escribir título
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x_start, current_y, "Descuentos:")
    current_y -= line_height + 5  # Reducir Y para el siguiente contenido
    # Cambiar la fuente para los detalles
    c.setFont("Helvetica", 7)
    # Condicionales y datos
    if data["procesado"]["descuento_por_diferencia_peso"] > 0:
        c.drawString(
            x_start,
            current_y,
            f"B/V 004-:        dscto. x dif. peso $ {data['procesado']['descuento_por_diferencia_peso']:.2f}"
        )
        current_y -= line_height

    if data["procesado"]["descuento_sacos_faltantes"] > 0:
        c.drawString(
            x_start,
            current_y,
            f"B/V 004-:        dscto. x sacos falt. $ {data['procesado']['descuento_sacos_faltantes']:.2f}"
        )
        current_y -= line_height

    if data["procesado"]["total_descuento_solo_sacos"] > 0:
        c.drawString(
            x_start,
            current_y,
            f"B/V 004-:        dscto. x sacos R,H,M. $ {data['procesado']['total_descuento_solo_sacos']:.2f}"
        )
        current_y -= line_height

    if data["procesado"]["total_descuento_estiba"] > 0:
        c.drawString(
            x_start,
            current_y,
            f"B/V 004-:        dscto. x estibaje $ {data['procesado']['total_descuento_estiba']:.2f}"
        )
        current_y -= line_height

    # Total
    c.setFont("Helvetica-Bold", 8)
    c.drawString(
        x_start,
        current_y - 5,
        f"TOTAL A DESCONTAR:   $ {data['procesado']['total_dsct']:.2f}"
    )

    # Guardar el PDF en el buffer
    c.showPage()
    c.save()

    # Mover el puntero del buffer al inicio para poder leerlo
    buffer.seek(0)

    # Crear la respuesta HTTP con el archivo PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="reporte_calculo_flete.pdf"'
    response.write(buffer.getvalue())
    buffer.close()

    return response

def generar_reporte_detallado(request, styleSheet=None):
    data_unprocess = json.loads(request.body)
    data = procesar_data_reporte(data_unprocess)
    buffer = io.BytesIO()

    # Crear un archivo PDF
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    c.setTitle(" Reporte cálculo de flete detallado")


    # Obtener las dimensiones de la página
    width, height = landscape(A4)



    # primera linea
    c.setFont("Helvetica", 10)
    current_y = height - 40
    c.drawString(40, current_y, f"{data['procesado']['empresa']}")

    current_datetime = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    c.setFont("Helvetica",7)
    c.drawString(width-130, current_y, f"{current_datetime}")

    #segunda linea
    c.setFont("Helvetica-Bold", 10)
    texto = f"{data['dataForm']['producto']}"
    text_width = c.stringWidth(texto, "Helvetica", 12)
    # Calcular la posición X para centrar el texto
    x_position = (width - text_width) / 2 + 20
    # Dibujar el texto centrado
    current_y-=20
    c.drawString(x_position, current_y, texto)

    # tercera linea
    current_y -= 20
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, current_y, f"CARTA PORTE: {data['dataForm']['cartaPorte']}")



    # CUARTA LINEA DATOS DEL DESPACHO
    ordenes_recojo = data['dataForm']['ordenRecojo']
    ordenes_string = " / ".join([f"{row['oc']['numero_oc']}({row['numeroRecojo']})" for row in ordenes_recojo])
    textos = [
        f"N° DUA: {data['dataForm']['dua']}",
        f"Fec. Num.: {data['procesado']['fecha_numeracion']}",
        f"Factura N°: {data['dataForm']['numFactura']}",
        f"OC: {ordenes_string}",
        f"CANT. {data['procesado']['total_sacos_cargados']} sacos",
        f" {data['dataForm']['pesoNetoCrt']} kg"
    ]
    # Fuente y tamaño
    font_name = "Helvetica"
    font_size = 8
    c.setFont(font_name, font_size)
    # Espacio entre columnas
    margin = 30
    # Posición inicial en Y (mantener fija)
    current_y -= 20
    y_position = current_y
    # Posición inicial en X
    x_position = 40
    # Dibujar todos los textos en una línea tomando en cuenta su tamaño para darle cierto espaciado
    for texto in textos:
        # Calcular el ancho de cada texto
        text_width = c.stringWidth(texto, font_name, font_size)

        # Dibujar el texto
        c.drawString(x_position, y_position, texto)

        # Actualizar la posición X para el siguiente texto
        x_position += text_width + margin

        # Si el texto excede el tamaño de la página en el eje X, detener la creación del contenido
        if x_position > width:
            break



    #QUINTA LINEA DONDE INICIA LA TABLA:
    current_y -= 20
    # Datos iniciales de la tabla #1 (Titulos y subtitulos)
    data_table = [
        # Línea de títulos
        ['FEC. INGRE.', 'EMPRESA DE TRANSP.', 'CARGA', '', '', '', 'DESCARGA', '', '', '', 'DESCUENTOS', '', '', '',
         ''],
        ['', '', 'N°', 'Placa S.', 'Sacos C.', 'Peso S.', 'Placa L.', 'Sacos D.', 'Peso L.', 'Merma', 'S. Falt.',
         'S. Rotos', 'S. Humed.', 'S. Mojad.', 'Estibaje'],
    ]
    # Agregamos datos dinamicamente a la tabla
    styles = getSampleStyleSheet()
    parrafo = Paragraph(f"<para align=center spaceb=3><b>{data['dataForm']['transportista']}</b></para>",
                      styles["BodyText"])
    for i, row in enumerate(data['dataTable']):
        if i == 0:  # Primera fila (datos especiales)
            data_table.append([
                f"{data['procesado']['fecha_numeracion']}",
                parrafo,
                str(row['numero']),
                row['placa'],
                str(row['sacosCargados']),
                str(row['pesoSalida']),
                row['placaLlegada'],
                str(row['sacosDescargados']),
                str(row['pesoLlegada']),
                str(row['merma']),
                str(row['sacosFaltantes']),
                str(row['sacosRotos']),
                str(row['sacosHumedos']),
                str(row['sacosMojados']),
                str(row['pagoEstiba']),
            ])
        else:  # Para el resto de las filas
            data_table.append([
                "",  # Cadena vacía
                "",  # Cadena vacía
                str(row['numero']),
                row['placa'],
                str(row['sacosCargados']),
                str(row['pesoSalida']),
                row['placaLlegada'],
                str(row['sacosDescargados']),
                str(row['pesoLlegada']),
                str(row['merma']),
                str(row['sacosFaltantes']),
                str(row['sacosRotos']),
                str(row['sacosHumedos']),
                str(row['sacosMojados']),
                str(row['pagoEstiba']),
            ])
    # Agregamos
    data_table.append(
        ['', f"Flete x TM: $ {data['dataForm']['fletePactado']:.2f}", f"{data['procesado']['len_tabla']}", 'TOTAL',
         f"{data['procesado']['total_sacos_cargados']}", f"{data['procesado']['suma_peso_salida']}", 'TOTAL',
         f"{data['procesado']['total_sacos_descargados']}", f"{data['procesado']['suma_peso_llegada']}",
         f"{data['procesado']['merma_total']}", f"{data['procesado']['total_sacos_faltantes']}",
         f"{data['procesado']['total_sacos_rotos']}", f"{data['procesado']['total_sacos_humedos']}",
         f"{data['procesado']['total_sacos_mojados']}"])
    # Ancho de columnas
    col_widths = [60, 110, 20, 60, 40, 60, 60, 60, 40, 60, 30, 30, 30, 30, 70]
    # Crear la tabla
    table = Table(data_table, colWidths=col_widths, rowHeights=15)
    color_rosa = Color(red=250 / 255, green=210 / 255, blue=202 / 255)
    color_celeste = Color(red=176 / 255, green=225 / 255, blue=250 / 255)
    numero_filas = len(data_table)
    # Estilos para la tabla
    table_style = TableStyle([
        ('SPAN', (0, 0), (0, 1)),  # Fusionar "FEC. INGRE."
        ('SPAN', (1, 0), (1, 1)),  # Fusionar "EMPRESA DE TRANSP."
        ('SPAN', (2, 0), (5, 0)),  # Fusionar "CARGA"
        ('SPAN', (6, 0), (9, 0)),  # Fusionar "DESCARGA"
        ('SPAN', (10, 0), (14, 0)),  # Fusionar "DESCUENTOS"
        ('SPAN', (0, 2), (0, numero_filas - 1)),  # Fusionar "fecha"
        ('SPAN', (1, 2), (1, numero_filas - 2)),  # Fusionar "nombre empresa transporte"
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Centrar todo
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Fondo gris para los títulos
        ('BACKGROUND', (0, 1), (-1, 1), colors.lightgrey),  # Fondo gris claro para subtítulos
        ('BACKGROUND', (2, 0), (5, numero_filas), colors.yellowgreen),  # Color "CARGA"
        ('BACKGROUND', (6, 0), (9, numero_filas), color_celeste),  # Color "DESCARGA"
        ('BACKGROUND', (10, 0), (14, numero_filas), color_rosa),  # Fusionar "DESCUENTOS"
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Negrita en títulos
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),  # Negrita en subtítulos

        # ('FONTSIZE', (15, 0), (15, numero_filas - 1), 5),
    ])
    table_style.add('FONTSIZE', (1, 2), (1, numero_filas - 2), 5)
    # Aplicar estilos a la tabla
    table.setStyle(table_style)
    # Calcular la altura total de la tabla
    table_height = len(data_table) * 15 # 20 es la altura de cada fila
    # Calcular la posición de la tabla en la página
    table.wrapOn(c, width, height)
    table.drawOn(c, 40, current_y - table_height )  # Ajustar la posición dependiendo de la cantidad de filas

    current_y= current_y - table_height - 20

    # Data para la tabla #2 (Detalles de pesos)
    second_data_table=[
        ["Faltante peso Sta. Cruz - Desaguadero:",f"{data['procesado']['diferencia_de_peso']:.2f} Kg."],
        ["Merma permitida:",f"{data['dataExtra']['mermaPermitida']:.2f} Kg."],
        ["Desc. diferencia de peso:",f"{data['procesado']['diferencia_peso_por_cobrar']:.2f} Kg."],
        ["Desc. sacos falantes:",f"{data['procesado']['descuento_peso_sacos_faltantes']:.2f} Kg."]
    ]
    second_col_widths=[110,50]
    second_table = Table(second_data_table, colWidths=second_col_widths,rowHeights=14)
    second_table_style=TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Centrar
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
    ])
    second_table.setStyle(second_table_style)
    second_table_height = len(second_data_table) * 14
    second_table.wrapOn(c, width, height)
    second_table.drawOn(c, 40,current_y - second_table_height )


    #Data para la tabla #3 (Resumen de descuentos)
    third_data_table=[
        ["FLETE", f"$ {data['procesado']['flete_base']:.2f}"],
        ["Dscto. por dif. de peso", f"$ {data['procesado']['descuento_por_diferencia_peso']:.2f}"],
        ["Dscto. por sacos faltantes", f"$ {data['procesado']['descuento_sacos_faltantes']:.2f}"],
        ["Dscto. por sacos rotos", f"$ {data['procesado']['descuento_sacos_rotos']:.2f}"],
        ["Dscto. por sacos humedos", f"$ {data['procesado']['descuento_sacos_humedos']:.2f}"],
        ["Dscto. por sacos mojados", f"$ {data['procesado']['descuento_sacos_mojados']:.2f}"],
        ["FLETE TOTAL", f"$ {data['procesado']['total_luego_dsctos_sacos']:.2f}"],
    ]
    #posicion en x para la tercera tabla:
    x_position_for_third_table=260
    third_col_widths = [110, 40]
    third_table = Table(third_data_table, colWidths=third_col_widths, rowHeights=14)
    third_table_style = TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
    ])
    third_table.setStyle(third_table_style)
    third_table_height = len(third_data_table) * 14
    third_table.wrapOn(c, width, height)
    third_table.drawOn(c, x_position_for_third_table, current_y - third_table_height )

    #PARA VALIDACION DEL VALOR CRT
    peso_total_toneladas=data['dataForm']['pesoNetoCrt']/1000
    flete_pactado= data["dataForm"]["fletePactado"]
    monto_flete = round((peso_total_toneladas * flete_pactado),2)
    flete_base=data["procesado"]["flete_base"]
    estado_crt=""
    if(flete_base <= monto_flete):
        estado_crt="NO SOBREPASA VALOR CRT"
    else:
        estado_crt="SOBREPASA VALOR CRT"
    c.setFont("Helvetica-Bold", 7)
    c.drawString(x_position_for_third_table + 150 +10 , current_y-8,estado_crt)



    # Primera LLave para descuento sacos RMH
    c.setLineWidth(0.5)
    first_x_position_for_key = x_position_for_third_table + 150 +10
    first_key_y_position = current_y - (14 * 3)
    last_key_y_position=first_key_y_position - (14 * 3)
    middle_of_key= (first_key_y_position + last_key_y_position)/2
    inclinacion=2
    c.line(first_x_position_for_key, first_key_y_position-inclinacion, first_x_position_for_key, last_key_y_position+inclinacion)
    c.line(first_x_position_for_key, first_key_y_position-inclinacion, first_x_position_for_key-inclinacion, first_key_y_position)
    c.line(first_x_position_for_key-inclinacion, last_key_y_position , first_x_position_for_key , last_key_y_position+inclinacion)
    c.line(first_x_position_for_key, middle_of_key, first_x_position_for_key+inclinacion,middle_of_key)
    c.setFont("Helvetica-Bold", 7)
    c.drawString(first_x_position_for_key+inclinacion+4,middle_of_key , f"$ {data['procesado']['total_descuento_solo_sacos']:.2f}")

    # Segunda llave para descuento de sacos RMH
    c.setLineWidth(0.5)

    # Coordenadas para dibujar segunda la llave
    key_start_x = first_x_position_for_key+inclinacion+4 + 40
    key_start_y = current_y - 14
    key_end_y = key_start_y - (14 * 5)
    key_middle_y = (key_start_y + key_end_y) / 2
    key_inclination = 2

    # Dibujar la llave
    c.line(key_start_x, key_start_y - key_inclination, key_start_x, key_end_y + key_inclination)
    c.line(key_start_x, key_start_y - key_inclination, key_start_x - key_inclination, key_start_y)
    c.line(key_start_x - key_inclination, key_end_y, key_start_x, key_end_y + key_inclination)
    c.line(key_start_x, key_middle_y, key_start_x + key_inclination, key_middle_y)

    # Texto del descuento
    c.setFont("Helvetica-Bold", 7)
    c.drawString(key_start_x + key_inclination + 4, key_middle_y,
                 f"$ {data['procesado']['aux_descuento']:.2f}")


    # Lista de placas y estado de estiba:
    y_position_for_list=current_y-third_table_height-10
    c.setFont("Helvetica", 6)
    for item in data["procesado"]["pago_estiba_list"]:
        # Crear texto en una línea
        c.drawString(x_position_for_third_table, y_position_for_list, f"{item['placa']} {item['detalle']}")  # Placa y detalle
        c.drawString(x_position_for_third_table + 130, y_position_for_list, f"$ {item['monto_descuento']:.2f}")  # Monto alineado a la derecha
        y_position_for_list -= 10  # Reducir la posición vertical para la próxima línea


    # Coordenadas para dibujar la tercera llave
    tirth_key_start_x = first_x_position_for_key
    tirth_key_start_y = current_y-third_table_height-5
    tirth_key_end_y = y_position_for_list +10
    tirth_key_middle_y = (tirth_key_start_y + tirth_key_end_y) / 2
    tirth_key_inclination = 2
    # Dibujar la tercera llave
    c.line(tirth_key_start_x, tirth_key_start_y - tirth_key_inclination, tirth_key_start_x,
           tirth_key_end_y + tirth_key_inclination)
    c.line(tirth_key_start_x, tirth_key_start_y - tirth_key_inclination, tirth_key_start_x - tirth_key_inclination,
           tirth_key_start_y)
    c.line(tirth_key_start_x - tirth_key_inclination, tirth_key_end_y, tirth_key_start_x,
           tirth_key_end_y + tirth_key_inclination)
    c.line(tirth_key_start_x, tirth_key_middle_y, tirth_key_start_x + tirth_key_inclination, tirth_key_middle_y)
    # Texto del descuento
    c.setFont("Helvetica-Bold", 7)
    c.drawString(tirth_key_start_x + tirth_key_inclination + 4, tirth_key_middle_y-2,
                 f"$ {data['procesado']['total_descuento_estiba']:.2f}")


    # Escribir linea de neto a pagar
    table_total_data = [
        ["TOTAL A PAGAR:", f"${data['procesado']['total_a_pagar']:.2f}"]
    ]
    # Crear la tabla
    table_total = Table(table_total_data, colWidths=[110, 40])  # Anchos de columnas
    # Aplicar estilo a la tabla
    table_total.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.orange),  # Fondo naranja
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # Texto en color negro
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alineación izquierda primera columna
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),  # Fuente en negrita
        ('FONTSIZE', (0, 0), (-1, -1), 8),  # Tamaño de la fuente
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'), # Alineación derecha segunda columna
    ]))
    altura_tabla_total=len(table_total_data) * 14
    table_total.wrapOn(c, width, height)
    table_total.drawOn(c, x_position_for_third_table, y_position_for_list - altura_tabla_total-20)


    # Escribir linea de comprobacion
    total_a_pagar=data["procesado"]["total_a_pagar"]
    total_desctos=data["procesado"]["total_dsct"]
    monto_final=total_a_pagar+total_desctos
    diferencia= round((monto_final - flete_base),2)
    table_total_data = [
        ["COMPROBACIÓN", f"$ {monto_final:.2f}",f"$ {diferencia}"]
    ]
    # Crear la tabla
    table_total = Table(table_total_data, colWidths=[80, 60,60])  # Anchos de columnas
    # Aplicar estilo a la tabla
    table_total.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.orange),  # Fondo naranja
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # Texto en color negro
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alineación izquierda primera columna
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),  # Fuente en negrita
        ('FONTSIZE', (0, 0), (-1, -1), 8),  # Tamaño de la fuente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
    ]))
    altura_tabla_total = len(table_total_data) * 14
    table_total.wrapOn(c, width, height)
    table_total.drawOn(c, 40, y_position_for_list - altura_tabla_total - 70)


    # Data para la tabla #4 (resumen de descuentos)
    new_x_position = 260 + 150 + 150
    fourth_data_table=[
        ["FLETE", f"$ {data['procesado']['flete_base']:.2f}"],
        ["Dscto.", f"$ {data['procesado']['total_dsct']:.2f}"],
        ["Neto.", f"$ {data['procesado']['total_a_pagar']:.2f}"],
    ]
    fourth_col_widths = [30, 50]
    fourth_table = Table(fourth_data_table, colWidths=fourth_col_widths, rowHeights=14)
    fourth_table_style = TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
    ])
    fourth_table.setStyle(fourth_table_style)
    fourth_table_height = len(fourth_data_table) * 14
    fourth_table.wrapOn(c, width, height)
    fourth_table.drawOn(c, new_x_position, current_y - fourth_table_height - (14*3))
    c.setFont("Helvetica-Bold", 7)
    c.drawString(new_x_position, current_y - fourth_table_height + 5, "RESUMEN")

    # Data para la tabla #5 (determinacion de costos)
    second_new_x_position = new_x_position + 120
    c.setFont("Helvetica-Bold", 6)
    c.drawString(second_new_x_position, current_y+5, "Determinacion de costo x Kg.")
    fifth_data_table=[
        ["C.P", f"{data['procesado']['precio_por_tonelada']:.2f}"],
        ["F.B", f"$ {data['dataForm']['fletePactado']:.2f}"],
        ["G.N", f"{data['dataExtra']['gastosNacionalizacion']}"],
        ["MF CIA", f"{data['dataExtra']['margenFinanciero']}"],
        ["IGV 18%", f"{data['procesado']['igv']}"],
        ["PRECIO DES", f"{data['procesado']['precio_bruto_final']}"],
        ["COSTO TM", f"{data['procesado']['precio_por_tonelada_final']}"],
        ["COSTO Kg", f"{data['procesado']['precio_por_kg_final']}"],
    ]
    fifth_col_widths = [50, 40]
    fifth_table = Table(fifth_data_table, colWidths=fifth_col_widths, rowHeights=14)
    fifth_table_style = TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
    ])
    fifth_table.setStyle(fifth_table_style)
    fifth_table_height = len(fifth_data_table) * 14
    fifth_table.wrapOn(c, width, height)
    fifth_table.drawOn(c, second_new_x_position, current_y - fifth_table_height)


    # SALTO DE LINEA PARA ESPACIO DE DESCUENTOS
    current_y = current_y - second_table_height - 20
    x_start = 40  # Eje X constante
    line_height = 10  # Interlineado
    # Escribir título
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x_start, current_y, "Descuentos:")
    current_y -= line_height + 5  # Reducir Y para el siguiente contenido
    # Cambiar la fuente para los detalles
    c.setFont("Helvetica", 7)
    # Condicionales y datos
    if data["procesado"]["descuento_por_diferencia_peso"] > 0:
        c.drawString(
            x_start,
            current_y,
            f"B/V 004-:        dscto. x dif. peso $ {data['procesado']['descuento_por_diferencia_peso']:.2f}"
        )
        current_y -= line_height

    if data["procesado"]["descuento_sacos_faltantes"] > 0:
        c.drawString(
            x_start,
            current_y,
            f"B/V 004-:        dscto. x sacos falt. $ {data['procesado']['descuento_sacos_faltantes']:.2f}"
        )
        current_y -= line_height

    if data["procesado"]["total_descuento_solo_sacos"] > 0:
        c.drawString(
            x_start,
            current_y,
            f"B/V 004-:        dscto. x sacos R,H,M. $ {data['procesado']['total_descuento_solo_sacos']:.2f}"
        )
        current_y -= line_height

    if data["procesado"]["total_descuento_estiba"] > 0:
        c.drawString(
            x_start,
            current_y,
            f"B/V 004-:        dscto. x estibaje $ {data['procesado']['total_descuento_estiba']:.2f}"
        )
        current_y -= line_height

    # Total
    c.setFont("Helvetica-Bold", 8)
    c.drawString(
        x_start,
        current_y - 5,
        f"TOTAL A DESCONTAR:   $ {data['procesado']['total_dsct']:.2f}"
    )





    # Guardar el PDF en el buffer
    c.showPage()
    c.save()

    # Mover el puntero del buffer al inicio para poder leerlo
    buffer.seek(0)

    # Crear la respuesta HTTP con el archivo PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="reporte_calculo_flete.pdf"'
    response.write(buffer.getvalue())
    buffer.close()

    return response

def registrar_despacho(request):
    if request.method == 'POST':
        try:
            # Parsear el JSON de la request
            data = json.loads(request.body)
            data_procesado = procesar_data_reporte(data)
            # Procesar `dataForm`
            data_form = data.get('dataForm', {})
            empresa, _ = Empresa.objects.get_or_create(nombre_empresa=data_form.get('empresa'))
            providers, _ = ProveedorTransporte.objects.get_or_create(nombre_proveedor=data_form.get('proveedor'))
            transportista, _ = Transportista.objects.get_or_create(nombre_transportista=data_form.get('transportista'))
            fecha_llegada = data.get('dataExtraForm', {}).get('fechaLlegada', None)
            # Iniciar transacción
            with transaction.atomic():
                # Crear el Despacho
                despacho = Despacho.objects.create(
                    proveedor=providers,
                    dua=data_form.get('dua', ''),
                    fecha_numeracion=data_form.get('fechaNumeracion'),
                    carta_porte=data_form.get('cartaPorte', None),
                    num_factura=data_form.get('numFactura', ''),
                    transportista=transportista,
                    flete_pactado=data_form.get('fletePactado', 0.0),
                    peso_neto_crt=data_form.get('pesoNetoCrt', 0.0),
                    fecha_llegada=fecha_llegada
                )

                pdf_bytes = generar_reporte_pdf(data_procesado)
                # Luego, asignarlo al campo FileField de tu modelo
                despacho.archivo_pdf = pdf_bytes
                # Guardar el modelo
                despacho.save()




                # Procesar cada orden de compra y su número de recojo
                for orden_recojo in data_form.get('ordenRecojo', []):
                    oc_data = orden_recojo.get('oc', {})
                    numero_recojo = orden_recojo.get('numeroRecojo')
                    peso_asignado = orden_recojo.get('peso_asignado',0)

                    # Validar datos de `oc_data`
                    if not all(k in oc_data for k in ['codigo_producto', 'producto', 'proveedor', 'numero_oc', 'precio_unitario', 'cantidad']):
                        raise ValueError("Faltan datos necesarios en `oc`.")

                    # Crear o recuperar el producto
                    producto, _ = Producto.objects.get_or_create(
                        codigo_producto=oc_data['codigo_producto'],
                        nombre_producto=oc_data['producto'],
                        proveedor_marca=oc_data['proveedor']
                    )

                    # Crear o recuperar la OrdenCompra
                    print(f'este es el precio que se asignara a la oc : ',oc_data['precio_unitario'])
                    orden_compra, _ = OrdenCompra.objects.get_or_create(
                        empresa=empresa,
                        numero_oc=oc_data['numero_oc'],
                        producto=producto,
                        defaults={
                            'precio_producto': oc_data['precio_unitario'],
                            'cantidad': oc_data['cantidad']
                        }
                    )

                    # Verificar si ya existe la combinación de OC y número de recojo
                    if OrdenCompraDespacho.objects.filter(
                        orden_compra=orden_compra,
                        numero_recojo=numero_recojo
                    ).exists():
                        return JsonResponse(
                            {
                                'status': 'error',
                                'message': f'La OC "{orden_compra.numero_oc}" con número de recojo "{numero_recojo}" ya existe.'
                            },
                            status=400
                        )

                    # Crear la relación OrdenCompraDespacho
                    OrdenCompraDespacho.objects.create(
                        despacho=despacho,
                        orden_compra=orden_compra,
                        cantidad_asignada=peso_asignado,
                        numero_recojo=numero_recojo
                    )

                # Procesar `dataTable` para DetalleDespacho
                for item in data.get('dataTable', []):
                    DetalleDespacho.objects.create(
                        despacho=despacho,
                        placa_salida=item.get('placa', ''),
                        sacos_cargados=item.get('sacosCargados', 0),
                        peso_salida=item.get('pesoSalida', 0.0),
                        placa_llegada=item.get('placaLlegada', ''),
                        sacos_descargados=item.get('sacosDescargados', 0),
                        peso_llegada=item.get('pesoLlegada', 0.0),
                        merma=item.get('merma', 0),
                        sacos_faltantes=item.get('sacosFaltantes', 0),
                        sacos_rotos=item.get('sacosRotos', 0),
                        sacos_humedos=item.get('sacosHumedos', 0),
                        sacos_mojados=item.get('sacosMojados', 0),
                        pago_estiba=item.get('pagoEstiba', 0.0),
                        cant_desc=item.get('cantDesc', 0)
                    )

                # Procesar `dataExtraForm` para ConfiguracionDespacho
                data_extra_form = data.get('dataExtraForm', {})
                ConfiguracionDespacho.objects.create(
                    despacho=despacho,
                    merma_permitida=data_extra_form.get('mermaPermitida', 0.0),
                    precio_prod=data_extra_form.get('precioProd', 0.0),
                    gastos_nacionalizacion=data_extra_form.get('gastosNacionalizacion', 0.0),
                    margen_financiero=data_extra_form.get('margenFinanciero', 0.0),
                    precio_sacos_rotos=data_extra_form.get('precioSacosRotos', 0.0),
                    precio_sacos_humedos=data_extra_form.get('precioSacosHumedos', 0.0),
                    precio_sacos_mojados=data_extra_form.get('precioSacosMojados', 0.0),
                    tipo_cambio_desc_ext=data_extra_form.get('tipoCambioDescExt', 0.0),
                    precio_estiba=data_extra_form.get('precioEstibaTonelada', 0.0),
                )

                for item in data_extra_form.get('otrosGastos', []):
                    #(item)
                    GastosExtra.objects.create(
                        despacho=despacho,
                        descripcion=item['descripcion'],
                        monto=item['monto']
                    )

            return JsonResponse({'status': 'success', 'message': 'Registro realizado correctamente'}, status=201)


        except IntegrityError as e:

            return JsonResponse({
                'status': 'error',
                'message': f'Error de integridad: {str(e)}. Asegúrese de que los datos sean correctos.'
            }, status=400)
        except ValueError as ve:
            return JsonResponse({'status': 'error', 'message1': str(ve)}, status=400)
        except Exception as e:
            error_trace = traceback.format_exc()
            return JsonResponse({'status': 'error', 'message2': str(e), 'trace': error_trace}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

def listar_estiba(request):
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

    try:
        # Obtener y limpiar las fechas
        fecha_inicio = request.GET.get('fecha_inicio').strip()
        fecha_fin = request.GET.get('fecha_fin').strip()
        empresa = request.GET.get('empresa')

        # Validar formato de fecha
        try:
            fecha_inicio = make_aware(datetime.strptime(fecha_inicio, "%Y-%m-%d"))
            fecha_fin = make_aware(datetime.strptime(fecha_fin, "%Y-%m-%d"))
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Formato de fecha inválido. Debe ser YYYY-MM-DD.'}, status=400)

        # Consulta en ORM
        resultados = list(DetalleDespacho.objects.filter(
            Q(pago_estiba="No pago estiba") | Q(pago_estiba="Pago parcial"),
            Q(despacho__fecha_llegada__isnull=False) & Q(despacho__fecha_llegada__range=(fecha_inicio, fecha_fin)),
            Q(despacho__ordenes_compra__empresa__nombre_empresa=empresa)
        ).select_related(
            'despacho',
            'despacho__configuraciondespacho',
            'despacho__transportista'
        ).values(
            'id',
            'pago_estiba',
            'despacho__fecha_llegada',
            'despacho__dua',
            'placa_llegada',
            'sacos_descargados',
            'cant_desc',
            'despacho__configuraciondespacho__tipo_cambio_desc_ext',
            'despacho__configuraciondespacho__precio_estiba',
            'despacho__transportista__nombre_transportista',
            'despacho__ordenes_compra__empresa__nombre_empresa'  # Incluir el nombre de la empresa en los resultados
        ).distinct())  # Agregar .distinct() para evitar duplicados

        # Agregar cálculo de `total_a_pagar`
        for row in resultados:
            # Usa 4 como valor por defecto
            precio_estiba = row.get('despacho__configuraciondespacho__precio_estiba')
            precio_estiba = float(precio_estiba) if precio_estiba is not None else 4

            if row['pago_estiba'] == "No pago estiba":
                total_a_pagar = (row['sacos_descargados'] * 50 / 1000) * precio_estiba
                row['sacos_pendientes_de_pago'] = row['sacos_descargados']
            elif row['pago_estiba'] == "Pago parcial":
                total_a_pagar = (row['cant_desc'] * 50 / 1000) * precio_estiba
                row['sacos_pendientes_de_pago'] = row['cant_desc']
            else:
                total_a_pagar = 0  # Si no coincide con ninguna condición

            # Agregar siempre el total a pagar
            row['total_a_pagar'] = f"S/ {total_a_pagar:.2f}"
            row['precio_estiba'] = row['despacho__configuraciondespacho__precio_estiba']

        return JsonResponse({'status': 'success', 'data': resultados}, status=200)

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def listar_despachos(request):
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

    try:
        # Ordenamiento
        sort_field = request.GET.get('sortField', 'fecha_numeracion')
        sort_order = request.GET.get('sortOrder', 'descend')

        if sort_order == 'descend':
            orden = f"-{sort_field}"
        else:
            orden = f"{sort_field}"

        # Query base
        despachos = Despacho.objects.select_related(
            'proveedor', 'transportista'
        ).prefetch_related(
            'ordenes_despacho__orden_compra'
        ).all()

        # 🔹 Filtro por DUA
        dua = request.GET.get('dua')
        if dua:
            despachos = despachos.filter(dua__icontains=dua)

        # Aplicar orden
        despachos = despachos.order_by(orden)

        # Paginación
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))
        paginator = Paginator(despachos, page_size)

        try:
            despachos_paginados = paginator.page(page)
        except PageNotAnInteger:
            despachos_paginados = paginator.page(1)
        except EmptyPage:
            despachos_paginados = paginator.page(paginator.num_pages)

        # Respuesta
        data = [
            {
                "id": despacho.id,
                "dua": despacho.dua,
                "fecha_numeracion": despacho.fecha_numeracion.strftime("%Y-%m-%d %H:%M:%S"),
                "carta_porte": despacho.carta_porte,
                "num_factura": despacho.num_factura,
                "flete_pactado": f"$ {despacho.flete_pactado:.2f}",
                "peso_neto_crt": float(despacho.peso_neto_crt),
                "fecha_llegada": despacho.fecha_llegada.strftime(
                    "%Y-%m-%d %H:%M:%S") if despacho.fecha_llegada else None,
                "proveedor_nombre": despacho.proveedor.nombre_proveedor,
                "transportista_nombre": despacho.transportista.nombre_transportista,
                "ordenes_compra": [
                    {
                        "numero_oc": oc.orden_compra.numero_oc,
                        "producto": oc.orden_compra.producto.nombre_producto,
                        "precio_producto": f"{oc.orden_compra.precio_producto:.2f}",
                        "cantidad": oc.orden_compra.cantidad,
                        "numero_recojo": oc.numero_recojo,
                        "cantidad_asignada": oc.cantidad_asignada
                    }
                    for oc in despacho.ordenes_despacho.all()
                ]
            }
            for despacho in despachos_paginados  # 🔹 aquí solo usamos los paginados
        ]

        return JsonResponse({
            'status': 'success',
            'data': data,
            'total_count': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': despachos_paginados.number if despachos_paginados else 1
        }, status=200)

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def listar_data_despacho(request):
    try:
        id_despacho = request.GET.get('id')

        if not id_despacho:  # Si no se proporciona un ID
            return JsonResponse({"error": "Falta el parámetro 'id'"}, status=400)

        # Buscar el despacho en la BD
        data = Despacho.objects.filter(id=id_despacho)  # Convertir QuerySet a lista
        #serializer=DespachoSerializer(data,many=True)
        #json_data=serializer.data
        #data_process=procesar_data_bd_reporte(json_data)

        if not data:  # Si la lista está vacía
            return JsonResponse({"error": "No se encontraron datos"}, status=404)

        return JsonResponse({"status":"success","data":data},safe=False)  # Retorna la lista de resultados

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def generar_reporte_pdf(data):

    buffer = io.BytesIO()  # Crear un buffer en memoria

    # Crear el PDF
    c = canvas.Canvas(buffer, pagesize=landscape(A4))

    c.setTitle(" Reporte cálculo de flete")

    # Obtener las dimensiones de la página
    width, height = landscape(A4)

    # primera linea
    c.setFont("Helvetica", 10)
    current_y = height - 40
    c.drawString(40, current_y, f"{data['procesado']['empresa']}")

    current_datetime = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    c.setFont("Helvetica", 7)
    c.drawString(width - 130, current_y, f"{current_datetime}")

    # segunda linea
    c.setFont("Helvetica-Bold", 10)
    texto = f"{data['dataForm']['producto']}"
    text_width = c.stringWidth(texto, "Helvetica", 12)
    # Calcular la posición X para centrar el texto
    x_position = (width - text_width) / 2 + 20
    # Dibujar el texto centrado
    current_y -= 20
    c.drawString(x_position, current_y, texto)

    # tercera linea
    current_y -= 20
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, current_y, f"CARTA PORTE: {data['dataForm']['cartaPorte']}")

    # CUARTA LINEA DATOS DEL DESPACHO
    ordenes_recojo = data['dataForm']['ordenRecojo']
    ordenes_string = " / ".join([f"{row['oc']['numero_oc']}({row['numeroRecojo']})" for row in ordenes_recojo])
    textos = [
        f"N° DUA: {data['dataForm']['dua']}",
        f"Fec. Num.: {data['procesado']['fecha_numeracion']}",
        f"Factura N°: {data['dataForm']['numFactura']}",
        f"OC: {ordenes_string}",
        f"CANT. {data['procesado']['total_sacos_cargados']} sacos",
        f" {data['dataForm']['pesoNetoCrt']} kg"
    ]
    # Fuente y tamaño
    font_name = "Helvetica"
    font_size = 8
    c.setFont(font_name, font_size)
    # Espacio entre columnas
    margin = 30
    # Posición inicial en Y (mantener fija)
    current_y -= 20
    y_position = current_y
    # Posición inicial en X
    x_position = 40
    # Dibujar todos los textos en una línea tomando en cuenta su tamaño para darle cierto espaciado
    for texto in textos:
        # Calcular el ancho de cada texto
        text_width = c.stringWidth(texto, font_name, font_size)

        # Dibujar el texto
        c.drawString(x_position, y_position, texto)

        # Actualizar la posición X para el siguiente texto
        x_position += text_width + margin

        # Si el texto excede el tamaño de la página en el eje X, detener la creación del contenido
        if x_position > width:
            break

    # QUINTA LINEA DONDE INICIA LA TABLA:
    current_y -= 20
    # Datos iniciales de la tabla #1 (Titulos y subtitulos)
    data_table = [
        # Línea de títulos
        ['FEC. INGRE.', 'EMPRESA DE TRANSP.', 'CARGA', '', '', '', 'DESCARGA', '', '', '', 'DESCUENTOS', '', '', '',
         ''],
        ['', '', 'N°', 'Placa S.', 'Sacos C.', 'Peso S.', 'Placa L.', 'Sacos D.', 'Peso L.', 'Merma', 'S. Falt.',
         'S. Rotos', 'S. Humed.', 'S. Mojad.', 'Estibaje'],
    ]
    styles = getSampleStyleSheet()
    parrafo = Paragraph(f"<para align=center spaceb=3><b>{data['dataForm']['transportista']}</b></para>",
                        styles["BodyText"])
    for i, row in enumerate(data['dataTable']):
        if i == 0:  # Primera fila (datos especiales)
            data_table.append([
                f"{data['procesado']['fecha_numeracion']}",
                parrafo,
                str(row['numero']),
                row['placa'],
                str(row['sacosCargados']),
                str(row['pesoSalida']),
                row['placaLlegada'],
                str(row['sacosDescargados']),
                str(row['pesoLlegada']),
                str(row['merma']),
                str(row['sacosFaltantes']),
                str(row['sacosRotos']),
                str(row['sacosHumedos']),
                str(row['sacosMojados']),
                str(row['pagoEstiba']),
            ])
        else:  # Para el resto de las filas
            data_table.append([
                "",  # Cadena vacía
                "",  # Cadena vacía
                str(row['numero']),
                row['placa'],
                str(row['sacosCargados']),
                str(row['pesoSalida']),
                row['placaLlegada'],
                str(row['sacosDescargados']),
                str(row['pesoLlegada']),
                str(row['merma']),
                str(row['sacosFaltantes']),
                str(row['sacosRotos']),
                str(row['sacosHumedos']),
                str(row['sacosMojados']),
                str(row['pagoEstiba']),
            ])
    # Agregamos fila de totales
    data_table.append(
        ['', f"Flete x TM: $ {data['dataForm']['fletePactado']:.2f}", f"{data['procesado']['len_tabla']}", 'TOTAL',
         f"{data['procesado']['total_sacos_cargados']}", f"{data['procesado']['suma_peso_salida']}", 'TOTAL',
         f"{data['procesado']['total_sacos_descargados']}", f"{data['procesado']['suma_peso_llegada']}",
         f"{data['procesado']['merma_total']}", f"{data['procesado']['total_sacos_faltantes']}",
         f"{data['procesado']['total_sacos_rotos']}", f"{data['procesado']['total_sacos_humedos']}",
         f"{data['procesado']['total_sacos_mojados']}"])
    # Ancho de columnas
    col_widths = [60, 110, 20, 60, 40, 60, 60, 60, 40, 60, 30, 30, 30, 30, 70]
    # Crear la tabla
    table = Table(data_table, colWidths=col_widths, rowHeights=15)
    color_rosa = Color(red=250 / 255, green=210 / 255, blue=202 / 255)
    color_celeste = Color(red=176 / 255, green=225 / 255, blue=250 / 255)
    numero_filas = len(data_table)
    # Estilos para la tabla
    table_style = TableStyle([
        ('SPAN', (0, 0), (0, 1)),  # Fusionar "FEC. INGRE."
        ('SPAN', (1, 0), (1, 1)),  # Fusionar "EMPRESA DE TRANSP."
        ('SPAN', (2, 0), (5, 0)),  # Fusionar "CARGA"
        ('SPAN', (6, 0), (9, 0)),  # Fusionar "DESCARGA"
        ('SPAN', (10, 0), (14, 0)),  # Fusionar "DESCUENTOS"
        ('SPAN', (0, 2), (0, numero_filas - 1)),  # Fusionar "fecha"
        ('SPAN', (1, 2), (1, numero_filas - 2)),  # Fusionar "nombre empresa transporte"
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Centrar todo
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Fondo gris para los títulos
        ('BACKGROUND', (0, 1), (-1, 1), colors.lightgrey),  # Fondo gris claro para subtítulos
        ('BACKGROUND', (2, 0), (5, numero_filas), colors.yellowgreen),  # Color "CARGA"
        ('BACKGROUND', (6, 0), (9, numero_filas), color_celeste),  # Color "DESCARGA"
        ('BACKGROUND', (10, 0), (14, numero_filas), color_rosa),  # Fusionar "DESCUENTOS"
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Negrita en títulos
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),  # Negrita en subtítulos
        # ('FONTSIZE', (15, 0), (15, numero_filas - 1), 5),
    ])
    # Aplicar estilos a la tabla
    table.setStyle(table_style)
    # Calcular la altura total de la tabla
    table_height = len(data_table) * 15  # 20 es la altura de cada fila
    # Calcular la posición de la tabla en la página
    table.wrapOn(c, width, height)
    table.drawOn(c, 40, current_y - table_height)  # Ajustar la posición dependiendo de la cantidad de filas

    current_y = current_y - table_height - 20

    # Data para la tabla #2 (Detalles de pesos)
    second_data_table = [
        ["Faltante peso Sta. Cruz - Desaguadero:", f"{data['procesado']['diferencia_de_peso']:.2f} Kg."],
        ["Merma permitida:", f"{data['dataExtra']['mermaPermitida']:.2f} Kg."],
        ["Desc. diferencia de peso:", f"{data['procesado']['diferencia_peso_por_cobrar']:.2f} Kg."],
        ["Desc. sacos falantes:", f"{data['procesado']['descuento_peso_sacos_faltantes']:.2f} Kg."]
    ]
    second_col_widths = [110, 50]
    second_table = Table(second_data_table, colWidths=second_col_widths, rowHeights=14)
    second_table_style = TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Centrar
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
    ])
    second_table.setStyle(second_table_style)
    second_table_height = len(second_data_table) * 14
    second_table.wrapOn(c, width, height)
    second_table.drawOn(c, 40, current_y - second_table_height)

    # Data para la tabla #3 (Resumen de descuentos)
    third_data_table = [
        ["FLETE", f"$ {data['procesado']['flete_base']:.2f}"],
        ["Dscto. por dif. de peso", f"$ {data['procesado']['descuento_por_diferencia_peso']:.2f}"],
        ["Dscto. por sacos faltantes", f"$ {data['procesado']['descuento_sacos_faltantes']:.2f}"],
        ["Dscto. por sacos rotos", f"$ {data['procesado']['descuento_sacos_rotos']:.2f}"],
        ["Dscto. por sacos humedos", f"$ {data['procesado']['descuento_sacos_humedos']:.2f}"],
        ["Dscto. por sacos mojados", f"$ {data['procesado']['descuento_sacos_mojados']:.2f}"],
        ["FLETE TOTAL", f"$ {data['procesado']['total_luego_dsctos_sacos']:.2f}"],
    ]
    # posicion en x para la tercera tabla:
    x_position_for_third_table = 260
    third_col_widths = [110, 40]
    third_table = Table(third_data_table, colWidths=third_col_widths, rowHeights=14)
    third_table_style = TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
    ])
    third_table.setStyle(third_table_style)
    third_table_height = len(third_data_table) * 14
    third_table.wrapOn(c, width, height)
    third_table.drawOn(c, x_position_for_third_table, current_y - third_table_height)

    # PARA VALIDACION DEL VALOR CRT
    peso_total_toneladas = data['dataForm']['pesoNetoCrt'] / 1000
    flete_pactado = data["dataForm"]["fletePactado"]
    monto_flete = round((peso_total_toneladas * flete_pactado), 2)
    flete_base = data["procesado"]["flete_base"]
    estado_crt = ""
    if (flete_base <= monto_flete):
        estado_crt = "NO SOBREPASA VALOR CRT"
    else:
        estado_crt = "SOBREPASA VALOR CRT"
    c.setFont("Helvetica-Bold", 7)
    c.drawString(x_position_for_third_table + 150 + 10, current_y - 8, estado_crt)

    # Primera LLave para descuento sacos RMH
    c.setLineWidth(0.5)
    first_x_position_for_key = x_position_for_third_table + 150 + 10
    first_key_y_position = current_y - (14 * 3)
    last_key_y_position = first_key_y_position - (14 * 3)
    middle_of_key = (first_key_y_position + last_key_y_position) / 2
    inclinacion = 2
    c.line(first_x_position_for_key, first_key_y_position - inclinacion, first_x_position_for_key,
           last_key_y_position + inclinacion)
    c.line(first_x_position_for_key, first_key_y_position - inclinacion, first_x_position_for_key - inclinacion,
           first_key_y_position)
    c.line(first_x_position_for_key - inclinacion, last_key_y_position, first_x_position_for_key,
           last_key_y_position + inclinacion)
    c.line(first_x_position_for_key, middle_of_key, first_x_position_for_key + inclinacion, middle_of_key)
    c.setFont("Helvetica-Bold", 7)
    c.drawString(first_x_position_for_key + inclinacion + 4, middle_of_key,
                 f"$ {data['procesado']['total_descuento_solo_sacos']:.2f}")

    # Segunda llave para descuento de sacos RMH
    c.setLineWidth(0.5)

    # Coordenadas para dibujar segunda la llave
    key_start_x = first_x_position_for_key + inclinacion + 4 + 40
    key_start_y = current_y - 14
    key_end_y = key_start_y - (14 * 5)
    key_middle_y = (key_start_y + key_end_y) / 2
    key_inclination = 2

    # Dibujar la llave
    c.line(key_start_x, key_start_y - key_inclination, key_start_x, key_end_y + key_inclination)
    c.line(key_start_x, key_start_y - key_inclination, key_start_x - key_inclination, key_start_y)
    c.line(key_start_x - key_inclination, key_end_y, key_start_x, key_end_y + key_inclination)
    c.line(key_start_x, key_middle_y, key_start_x + key_inclination, key_middle_y)

    # Texto del descuento
    c.setFont("Helvetica-Bold", 7)
    c.drawString(key_start_x + key_inclination + 4, key_middle_y,
                 f"$ {data['procesado']['aux_descuento']:.2f}")

    # Lista de placas y estado de estiba:
    y_position_for_list = current_y - third_table_height - 10
    c.setFont("Helvetica", 6)
    for item in data["procesado"]["pago_estiba_list"]:
        # Crear texto en una línea
        c.drawString(x_position_for_third_table, y_position_for_list,
                     f"{item['placa']} {item['detalle']}")  # Placa y detalle
        c.drawString(x_position_for_third_table + 130, y_position_for_list,
                     f"$ {item['monto_descuento']:.2f}")  # Monto alineado a la derecha
        y_position_for_list -= 10  # Reducir la posición vertical para la próxima línea

    for item in data["procesado"]["otros_gastos"]:
        c.drawString(x_position_for_third_table, y_position_for_list, f"{item['descripcion']}")
        c.drawString(x_position_for_third_table + 130, y_position_for_list, f"$ {item['monto']:.2f}")
        y_position_for_list -= 10

    # Coordenadas para dibujar la tercera llave
    tirth_key_start_x = first_x_position_for_key
    tirth_key_start_y = current_y - third_table_height - 5
    tirth_key_end_y = y_position_for_list + 10
    tirth_key_middle_y = (tirth_key_start_y + tirth_key_end_y) / 2
    tirth_key_inclination = 2
    # Dibujar la tercera llave
    c.line(tirth_key_start_x, tirth_key_start_y - tirth_key_inclination, tirth_key_start_x,
           tirth_key_end_y + tirth_key_inclination)
    c.line(tirth_key_start_x, tirth_key_start_y - tirth_key_inclination, tirth_key_start_x - tirth_key_inclination,
           tirth_key_start_y)
    c.line(tirth_key_start_x - tirth_key_inclination, tirth_key_end_y, tirth_key_start_x,
           tirth_key_end_y + tirth_key_inclination)
    c.line(tirth_key_start_x, tirth_key_middle_y, tirth_key_start_x + tirth_key_inclination, tirth_key_middle_y)
    # Texto del descuento
    total_decuento = data['procesado']['total_descuento_estiba'] + data['procesado']['total_otros_gastos']
    c.setFont("Helvetica-Bold", 7)
    c.drawString(tirth_key_start_x + tirth_key_inclination + 4, tirth_key_middle_y - 2,
                 f"$ {total_decuento:.2f}")

    # Escribir linea de neto a pagar
    table_total_data = [
        ["TOTAL A PAGAR:", f"${data['procesado']['total_a_pagar']:.2f}"]
    ]
    # Crear la tabla
    table_total = Table(table_total_data, colWidths=[110, 40])  # Anchos de columnas
    # Aplicar estilo a la tabla
    table_total.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.orange),  # Fondo naranja
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # Texto en color negro
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alineación izquierda primera columna
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),  # Fuente en negrita
        ('FONTSIZE', (0, 0), (-1, -1), 8),  # Tamaño de la fuente
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alineación derecha segunda columna
    ]))
    altura_tabla_total = len(table_total_data) * 14
    table_total.wrapOn(c, width, height)
    table_total.drawOn(c, x_position_for_third_table, y_position_for_list - altura_tabla_total - 20)

    # Data para la tabla #4 (resumen de descuentos)
    new_x_position = 260 + 150 + 150
    fourth_data_table = [
        ["FLETE", f"$ {data['procesado']['flete_base']:.2f}"],
        ["Dscto.", f"$ {data['procesado']['total_dsct']:.2f}"],
        ["Neto.", f"$ {data['procesado']['total_a_pagar']:.2f}"],
    ]
    fourth_col_widths = [30, 50]
    fourth_table = Table(fourth_data_table, colWidths=fourth_col_widths, rowHeights=14)
    fourth_table_style = TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
    ])
    fourth_table.setStyle(fourth_table_style)
    fourth_table_height = len(fourth_data_table) * 14
    fourth_table.wrapOn(c, width, height)
    fourth_table.drawOn(c, new_x_position, current_y - fourth_table_height - (14 * 3))
    c.setFont("Helvetica-Bold", 7)
    c.drawString(new_x_position, current_y - fourth_table_height + 5, "RESUMEN")

    # Data para la tabla #5 (determinacion de costos)
    second_new_x_position = new_x_position + 120
    c.setFont("Helvetica-Bold", 6)
    c.drawString(second_new_x_position, current_y + 5, "Determinacion de costo x Kg.")
    fifth_data_table = [
        ["C.P", f"{data['procesado']['precio_por_tonelada']:.2f}"],
        ["F.B", f"$ {data['dataForm']['fletePactado']:.2f}"],
        ["G.N", f"{data['dataExtra']['gastosNacionalizacion']}"],
        ["MF CIA", f"{data['dataExtra']['margenFinanciero']}"],
        ["IGV 18%", f"{data['procesado']['igv']}"],
        ["PRECIO DES", f"{data['procesado']['precio_bruto_final']}"],
        ["COSTO TM", f"{data['procesado']['precio_por_tonelada_final']}"],
        ["COSTO Kg", f"{data['procesado']['precio_por_kg_final']}"],
    ]
    fifth_col_widths = [50, 40]
    fifth_table = Table(fifth_data_table, colWidths=fifth_col_widths, rowHeights=14)
    fifth_table_style = TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
    ])
    fifth_table.setStyle(fifth_table_style)
    fifth_table_height = len(fifth_data_table) * 14
    fifth_table.wrapOn(c, width, height)
    fifth_table.drawOn(c, second_new_x_position, current_y - fifth_table_height)

    # SALTO DE LINEA PARA ESPACIO DE DESCUENTOS
    current_y = current_y - second_table_height - 20

    x_start = 40  # Eje X constante
    line_height = 10  # Interlineado
    # Escribir título
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x_start, current_y, "Descuentos:")
    current_y -= line_height + 5  # Reducir Y para el siguiente contenido
    # Cambiar la fuente para los detalles
    c.setFont("Helvetica", 7)
    # Condicionales y datos
    if data["procesado"]["descuento_por_diferencia_peso"] > 0:
        c.drawString(
            x_start,
            current_y,
            f"B/V 004-:        dscto. x dif. peso $ {data['procesado']['descuento_por_diferencia_peso']:.2f}"
        )
        current_y -= line_height

    if data["procesado"]["descuento_sacos_faltantes"] > 0:
        c.drawString(
            x_start,
            current_y,
            f"B/V 004-:        dscto. x sacos falt. $ {data['procesado']['descuento_sacos_faltantes']:.2f}"
        )
        current_y -= line_height

    if data["procesado"]["total_descuento_solo_sacos"] > 0:
        c.drawString(
            x_start,
            current_y,
            f"B/V 004-:        dscto. x sacos R,H,M. $ {data['procesado']['total_descuento_solo_sacos']:.2f}"
        )
        current_y -= line_height

    if data["procesado"]["total_descuento_estiba"] > 0:
        c.drawString(
            x_start,
            current_y,
            f"B/V 004-:        dscto. x estibaje $ {data['procesado']['total_descuento_estiba']:.2f}"
        )
        current_y -= line_height

    # Total
    c.setFont("Helvetica-Bold", 8)
    c.drawString(
        x_start,
        current_y - 5,
        f"TOTAL A DESCONTAR:   $ {data['procesado']['tota_dsct_sin_gastos_otros']:.2f}"
    )

    # Guardar el PDF en el buffer
    c.showPage()
    c.save()

    buffer.seek(0)  # Volver al inicio del buffer
    return buffer.getvalue()  # Devolver el PDF en binario

def descargar_pdf(request, despacho_id):
    despacho = get_object_or_404(Despacho, id=despacho_id)

    if despacho.archivo_pdf:
        response = HttpResponse(despacho.archivo_pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_despacho_{despacho.id}.pdf"'
        return response
    else:
        return HttpResponse("No hay PDF disponible para este despacho", status=404)

def generar_reporte_base_bd(request):

    id_despacho = request.GET.get('id')
    if not id_despacho:  # Si no se proporciona un ID
        return JsonResponse({"error": "Falta el parámetro 'id'"}, status=400)
    try:
        query=Despacho.objects.filter(id=id_despacho)
        data_un = DespachoSerializer(query.first()).data

        if not data_un:  # Si la lista está vacía
            return JsonResponse({"error": "No se encontraron datos"}, status=404)

        data = sanear_y_procesar_data(data_un)
        pdf=generar_reporte_pdf_con_data_bd(data)

        if not pdf:
            return JsonResponse({"error": "No se pudo generar el PDF"}, status=500)

        # Retornar el PDF como un archivo
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'
        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

#vistas para editar fletes
def obtener_data_flete_ant(request):
    id_despacho = request.GET.get('despacho_id')
    if not id_despacho:
        return JsonResponse({"error": "Falta el parámetro 'id'"}, status=400)

    try:
        despacho = Despacho.objects.filter(id=id_despacho).first()
        if not despacho:
            return JsonResponse({"error": "No se encontraron datos"}, status=404)

        despacho_serializado = DespachoSerializer(despacho).data

        response = {
            "dataForm": construir_data_form(despacho_serializado),
            "dataTable": construir_data_table(despacho_serializado["detalle_despacho"]),
            "dataExtraForm": construir_data_extra(despacho_serializado["configuracion_despacho"], despacho_serializado)
        }
        return JsonResponse(response, status=200, safe=False)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def obtener_data_flete(request):
    id_despacho = request.GET.get('despacho_id')
    if not id_despacho:
        return JsonResponse({"error": "Falta el parámetro 'id'"}, status=400)

    try:
        despacho = Despacho.objects.get(id=id_despacho)

        data = {
            "general": despacho,
            "detalle_despacho": despacho.detalledespacho_set.all(),
            "configuracion_despacho": despacho.configuraciondespacho_set.all(),
            "gastos_extra": despacho.gastosextra_set.all(),
        }

        serializer = DespachoCompletoSerializer(data)
        return JsonResponse(serializer.data, safe=False, status=200)

    except Despacho.DoesNotExist:
        return JsonResponse({"error": "No se encontró el despacho"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@api_view(['PUT'])
def actualizar_despacho(request, id):
    try:
        despacho = Despacho.objects.get(id=id)
    except Despacho.DoesNotExist:
        return Response({"error": "Despacho no encontrado"}, status=404)

    data = request.data

    # Actualizar campos del despacho
    general_serializer = DespachoWriteSerializer(despacho, data=data.get("general"), partial=True)
    if not general_serializer.is_valid():
        return Response({"errors": general_serializer.errors}, status=400)
    general_serializer.save()

    # Limpiar y recrear ordenes_despacho
    despacho.ordenes_despacho.all().delete()
    ordenes_despacho_data = data.get("ordenes", {}).get("ordenes_despacho", [])
    for orden in ordenes_despacho_data:
        orden['despacho'] = despacho.id
    orden_serializer = OrdenCompraDespachoWriteSerializer(data=ordenes_despacho_data, many=True)
    if orden_serializer.is_valid():
        orden_serializer.save()
    else:
        return Response({"errors": orden_serializer.errors}, status=400)

    # Limpiar y recrear detalle despacho
    despacho.detalledespacho_set.all().delete()
    detalle_data = data.get("detalle_despacho", [])
    for d in detalle_data:
        d['despacho'] = despacho.id
    detalle_serializer = DetalleDespachoWriteSerializer(data=detalle_data, many=True)
    if detalle_serializer.is_valid():
        detalle_serializer.save()
    else:
        return Response({"errors": detalle_serializer.errors}, status=400)

    # Limpiar y recrear configuraciones
    despacho.configuraciondespacho_set.all().delete()
    config_data = data.get("configuracion_despacho", [])
    for c in config_data:
        c['despacho'] = despacho.id
    config_serializer = ConfiguracionDespachoWriteSerializer(data=config_data, many=True)
    if config_serializer.is_valid():
        config_serializer.save()
    else:
        return Response({"errors": config_serializer.errors}, status=400)

    # Limpiar y recrear gastos extra
    despacho.gastosextra_set.all().delete()
    gastos_data = data.get("gastos_extra", [])
    for g in gastos_data:
        g['despacho'] = despacho.id
    gastos_serializer = GastosExtraWriteSerializer(data=gastos_data, many=True)
    if gastos_serializer.is_valid():
        gastos_serializer.save()
    else:
        return Response({"errors": gastos_serializer.errors}, status=400)

    return Response({"message": "Despacho actualizado correctamente"}, status=200)


@api_view(['GET', 'PUT'])
def despacho_editar(request, pk):
    try:
        despacho = Despacho.objects.get(pk=pk)
    except Despacho.DoesNotExist:
        return Response({"error": "Despacho no encontrado"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = DespachoEditarSerializer(despacho)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = EditarDespachoSerializer(despacho, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class CrearProveedorView(generics.CreateAPIView):
    queryset = ProveedorTransporte.objects.all()
    serializer_class = ProveedorTransporteSerializer

class CrearTransportistaView(generics.CreateAPIView):
    queryset = Transportista.objects.all()
    serializer_class = TransportistaSerializer

@api_view(['GET'])
def buscar_proveedores(request):
    query = request.GET.get('query', '')
    proveedores = ProveedorTransporte.objects.filter(
        Q(nombre_proveedor__icontains=query)
    )[:10]  # Limitamos resultados
    serializer = ProveedorTransporteSerializer(proveedores, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def buscar_transportistas(request):
    query = request.GET.get('query', '')
    transportistas = Transportista.objects.filter(
        Q(nombre_transportista__icontains=query)
    )[:10]  # Limitamos resultados
    serializer = TransportistaSerializer(transportistas, many=True)
    return Response(serializer.data)


# Listar detalles por despacho_id
@api_view(['GET'])
def obtener_detalles_despacho(request):
    despacho_id = request.GET.get('despacho_id')
    if not despacho_id:
        return Response({'error': 'Falta el parámetro despacho_id'}, status=400)

    detalles = DetalleDespacho.objects.filter(despacho_id=despacho_id)
    serializer = DetalleDespachoEditarSerializer(detalles, many=True)
    return Response(serializer.data)


# Crear nuevo detalle
@api_view(['POST'])
def crear_detalle_despacho(request):
    serializer = DetalleDespachoEditarSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


# Editar detalle
@api_view(['PUT'])
def editar_detalle_despacho(request, pk):
    try:
        detalle = DetalleDespacho.objects.get(pk=pk)
    except DetalleDespacho.DoesNotExist:
        return Response({'error': 'Detalle no encontrado'}, status=404)

    serializer = DetalleDespachoEditarSerializer(detalle, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=400)

# Eliminar detalle
@api_view(['DELETE'])
def eliminar_detalle_despacho(request, pk):
    try:
        detalle = DetalleDespacho.objects.get(pk=pk)
    except DetalleDespacho.DoesNotExist:
        return Response({'error': 'Detalle no encontrado'}, status=404)

    detalle.delete()
    return Response({'mensaje': 'Detalle eliminado'}, status=204)

@api_view(['GET'])
def obtener_configuracion_despacho(request):
    despacho_id = request.GET.get('despacho_id')
    if not despacho_id:
        return Response({'error': "Falta el parámetro 'despacho_id'"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        configuracion = ConfiguracionDespacho.objects.get(despacho_id=despacho_id)
        serializer = ConfiguracionDespachoEditarSerializer(configuracion)
        return Response(serializer.data)
    except ConfiguracionDespacho.DoesNotExist:
        return Response({'error': 'No existe configuración para ese despacho'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['PUT'])
def editar_configuracion_despacho(request, pk):
    try:
        configuracion = ConfiguracionDespacho.objects.get(pk=pk)
    except ConfiguracionDespacho.DoesNotExist:
        return Response({'error': 'Configuración no encontrada'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ConfiguracionDespachoEditarSerializer(configuracion, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def listar_gastos_extra(request):
    despacho_id = request.GET.get('despacho_id')
    if not despacho_id:
        return Response({'error': 'Se requiere el parámetro despacho_id'}, status=400)

    gastos = GastosExtra.objects.filter(despacho_id=despacho_id).order_by('-fecha_de_creacion')
    serializer = GastosExtraSerializer(gastos, many=True)
    return Response(serializer.data)

@api_view(['POST'])
def crear_gasto_extra(request):
    serializer = GastosExtraSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)

@api_view(['PUT', 'PATCH'])
def editar_gasto_extra(request, pk):
    try:
        gasto = GastosExtra.objects.get(pk=pk)
    except GastosExtra.DoesNotExist:
        return Response({'error': 'Gasto extra no encontrado'}, status=404)

    serializer = GastosExtraSerializer(gasto, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=400)

@api_view(['DELETE'])
def eliminar_gasto_extra(request, pk):
    try:
        gasto = GastosExtra.objects.get(pk=pk)
    except GastosExtra.DoesNotExist:
        return Response({'error': 'Gasto extra no encontrado'}, status=404)

    gasto.delete()
    return Response({'mensaje': 'Gasto eliminado correctamente'}, status=204)

class DespachoOrdenListCreateView(generics.ListCreateAPIView):
    """
    GET: Lista todas las órdenes asociadas a un despacho
    POST: Crea una relación OrdenCompraDespacho + OC + Producto si es necesario
    """
    serializer_class = OrdenDespachoRowSerializer

    def get_queryset(self):
        despacho_id = self.kwargs['despacho_id']
        return OrdenCompraDespacho.objects.filter(
            despacho_id=despacho_id
        ).select_related('orden_compra__empresa', 'orden_compra__producto')

    def perform_create(self, serializer):
        despacho_id = self.kwargs['despacho_id']
        despacho = get_object_or_404(Despacho, pk=despacho_id)
        serializer.save(despacho=despacho)

class DespachoOrdenRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET: Obtiene detalles combinados de una fila de despacho
    PUT/PATCH: Actualiza datos de la fila y de la orden de compra relacionada
    DELETE: Elimina la relación OrdenCompraDespacho (no la orden en sí)
    """
    serializer_class = OrdenDespachoRowSerializer
    lookup_url_kwarg = 'orden_despacho_id'

    def get_queryset(self):
        despacho_id = self.kwargs['despacho_id']
        return OrdenCompraDespacho.objects.filter(
            despacho_id=despacho_id
        ).select_related('orden_compra__empresa', 'orden_compra__producto')

class EmpresaListView(generics.ListAPIView):
    queryset = Empresa.objects.all().order_by('id')
    serializer_class = EmpresaEditarSerializer

class OrdenCompraDespachoViewSet(viewsets.ModelViewSet):
    queryset = OrdenCompraDespacho.objects.all()
    serializer_class = NewOrdenCompraDespachoSerializer

    def get_queryset(self):
        despacho_id = self.request.query_params.get('despacho_id')
        if despacho_id:
            return OrdenCompraDespacho.objects.filter(despacho_id=despacho_id)
        return super().get_queryset()

    def update(self, request, *args, **kwargs):
        despacho_id = request.data.get("despacho")
        instance = self.get_object()
        if despacho_id and str(instance.despacho_id) != str(despacho_id):
            return Response(
                {"detail": "El despacho no coincide con la relación a actualizar."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        despacho_id = request.data.get("despacho") or request.query_params.get("despacho_id")
        instance = self.get_object()
        if despacho_id and str(instance.despacho_id) != str(despacho_id):
            return Response(
                {"detail": "El despacho no coincide con la relación a eliminar."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        instance.delete()
        return Response({"detail": "Relación eliminada correctamente."}, status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['patch'], url_path="cambiar-orden")
    def cambiar_orden(self, request, pk=None):
        """
        Reemplaza la OrdenCompra en una relación, manteniendo cantidad_asignada y numero_recojo
        """
        instance = self.get_object()
        despacho_id = request.data.get("despacho")
        nueva_orden_id = request.data.get("orden_compra_id")

        if not despacho_id or not nueva_orden_id:
            return Response(
                {"detail": "Se requiere 'despacho' y 'orden_compra_id'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if str(instance.despacho_id) != str(despacho_id):
            return Response(
                {"detail": "El despacho no coincide con la relación a modificar."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            nueva_orden = OrdenCompra.objects.get(id=nueva_orden_id)
        except OrdenCompra.DoesNotExist:
            return Response({"detail": "La Orden de Compra no existe."}, status=status.HTTP_404_NOT_FOUND)

        # Validar que no se duplique (orden_compra + numero_recojo + despacho)
        if OrdenCompraDespacho.objects.filter(
            despacho_id=despacho_id,
            orden_compra=nueva_orden,
            numero_recojo=instance.numero_recojo
        ).exclude(id=instance.id).exists():
            return Response(
                {"detail": "Ya existe esta Orden de Compra con el mismo número de recojo para este despacho."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Actualizar solo la orden_compra
        instance.orden_compra = nueva_orden
        instance.save()

        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(["POST"])
def get_or_create_oc(request):
    numero_oc = request.data.get("numero_oc")
    defaults = {
        "empresa_id": request.data.get("empresa_id"),
        "producto_id": request.data.get("producto_id"),
        "precio_producto": request.data.get("precio_producto"),
        "cantidad": request.data.get("cantidad"),
    }

    oc, created = OrdenCompra.objects.get_or_create(
        numero_oc=numero_oc,
        defaults=defaults,
    )

    return Response(OrdenCompraSerializer(oc).data, status=status.HTTP_200_OK)

@csrf_exempt
@transaction.atomic
def crear_oc_desde_externa(request):
    """
    Crea una OrdenCompra (y sus entidades relacionadas)
    a partir de los datos devueltos por la base externa.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
        base_datos = data.get('base_datos')
        numero_oc = data.get('numero_oc')

        if not base_datos or not numero_oc:
            return JsonResponse({'error': 'Faltan parámetros requeridos'}, status=400)

        # --- 1️⃣ Consultar la BD externa usando tu misma lógica ---
        connection = get_db_connection(base_datos)
        with connection.cursor() as cursor:
            cursor.execute(f"""
                SELECT TOP 1 
                    i.CNUMERO,
                    i.CDESARTIC,
                    i.CCODARTIC,
                    i.NCANTIDAD,
                    i.CUNIDAD,
                    i.NPREUNITA,
                    i.NTOTVENT,
                    o.CDESPROVE,
                    o.CCODPROVE
                FROM IMPORD i
                INNER JOIN IMPORC o ON i.CNUMERO = o.CNUMERO
                WHERE i.CNUMERO = %s
            """, [numero_oc])
            row = cursor.fetchone()

        if not row:
            return JsonResponse({'error': 'No se encontró la OC en la base externa'}, status=404)

        # --- 2️⃣ Mapear los datos de la BD externa ---
        datos = {
            'numero_oc': row[0],
            'producto': row[1],
            'codigo_producto': row[2],
            'cantidad': row[3],
            'unidad_medida': row[4],
            'precio_unitario': row[5],
            'precio_total': row[6],
            'proveedor': row[7],
            'codprovee': row[8],
        }

        # --- 3️⃣ Crear o recuperar Empresa ---
        empresa, empresa_creada = Empresa.objects.get_or_create(
            nombre_empresa=datos['proveedor'],
            defaults={'ruc': datos['codprovee']}
        )

        # --- 4️⃣ Crear o recuperar Producto ---
        producto, producto_creado = Producto.objects.get_or_create(
            codigo_producto=datos['codigo_producto'],
            defaults={
                'nombre_producto': datos['producto'],
                'unidad_medida': datos['unidad_medida'],
                'proveedor_marca': datos['proveedor']
            }
        )

        # --- 5️⃣ Crear o recuperar OrdenCompra ---
        oc, oc_creada = OrdenCompra.objects.get_or_create(
            numero_oc=datos['numero_oc'],
            defaults={
                'empresa': empresa,
                'producto': producto,
                'cantidad': datos['cantidad'],
                'precio_producto': datos['precio_unitario']
            }
        )

        # --- 6️⃣ Respuesta estandarizada ---
        return JsonResponse({
            'orden_compra_id': oc.id,
            'numero_oc': oc.numero_oc,
            'empresa': empresa.nombre_empresa,
            'producto': producto.nombre_producto,
            'cantidad': oc.cantidad,
            'precio_producto': float(oc.precio_producto),
            'creada': oc_creada,
            'empresa_creada': empresa_creada,
            'producto_creado': producto_creado
        }, status=201 if oc_creada else 200)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

#aqui terminan las vistas para editar fletes

def exportar_reporte_estiba_excel(request):
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

    try:
        fecha_inicio = request.GET.get('fecha_inicio', '').strip()
        fecha_fin = request.GET.get('fecha_fin', '').strip()
        empresa_bd = request.GET.get('empresa', '').strip()

        # Mapeo a nombre comercial
        if empresa_bd == "bd_trading_starsoft":
            empresa = "TRADING SEMILLA SAC"
        elif empresa_bd == "bd_semilla_starsoft":
            empresa = "LA SEMILLA DE ORO SAC"
        elif empresa_bd == "bd_maxi_starsoft":
            empresa = "MAXIMILIAN INVERSIONES SA"
        else:
            empresa = empresa_bd  # fallback

        try:
            fecha_inicio_dt = make_aware(datetime.strptime(fecha_inicio, "%Y-%m-%d"))
            fecha_fin_dt = make_aware(datetime.strptime(fecha_fin, "%Y-%m-%d"))
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Formato de fecha inválido. Debe ser YYYY-MM-DD.'}, status=400)

        resultados = list(DetalleDespacho.objects.filter(
            Q(pago_estiba="No pago estiba") | Q(pago_estiba="Pago parcial"),
            Q(despacho__fecha_llegada__isnull=False) & Q(despacho__fecha_llegada__range=(fecha_inicio_dt, fecha_fin_dt)),
            Q(despacho__ordenes_compra__empresa__nombre_empresa=empresa_bd)
        ).select_related(
            'despacho',
            'despacho__transportista',
            'despacho__configuraciondespacho'
        ).values(
            'id',
            'pago_estiba',
            'despacho__fecha_llegada',
            'despacho__dua',
            'placa_llegada',
            'sacos_descargados',
            'cant_desc',
            'despacho__configuraciondespacho__precio_estiba',
            'despacho__transportista__nombre_transportista',
            'despacho__ordenes_compra__empresa__nombre_empresa'
        ).distinct())

        total_general = 0
        for row in resultados:
            # Usa 4 como valor por defecto
            precio_estiba = row['despacho__configuraciondespacho__precio_estiba']
            precio_estiba = float(precio_estiba) if precio_estiba is not None else 4
            print('precio estiba: ',precio_estiba)

            if row['pago_estiba'] == "No pago estiba":
                total_a_pagar = (row['sacos_descargados'] * 50 / 1000) * precio_estiba
                row['sacos_pendientes_de_pago'] = row['sacos_descargados']
            elif row['pago_estiba'] == "Pago parcial":
                total_a_pagar = (row['cant_desc'] * 50 / 1000) * precio_estiba
                row['sacos_pendientes_de_pago'] = row['cant_desc']
            else:
                total_a_pagar = 0

            row['total_a_pagar'] = total_a_pagar
            total_general += total_a_pagar

        # Crear Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Estiba"

        from openpyxl.styles import Font, Alignment

        # Primera fila: nombre de empresa (combinada)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        ws['A1'] = empresa.upper()
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Segunda fila: espacio
        ws.append([""] * 11)

        # Tercera fila: rango de fechas
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11)
        ws['A3'] = f"REPORTE DE ESTIBAJE DEL: {fecha_inicio} HASTA {fecha_fin}"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A3'].alignment = Alignment(horizontal='center')

        # Cuarta fila: cabeceras de tabla
        headers = [
            "ID", "Pago Estiba", "Fecha Llegada", "DUA", "Placa", "Sacos Descargados", "Cant. Desc.",
            "Transportista", "Empresa", "Sacos Pendientes de Pago", "Total a Pagar"
        ]
        ws.append(headers)

        for row in resultados:
            ws.append([
                row['id'],
                row['pago_estiba'],
                row['despacho__fecha_llegada'].strftime('%Y-%m-%d') if row['despacho__fecha_llegada'] else '',
                row['despacho__dua'],
                row['placa_llegada'],
                row['sacos_descargados'],
                row['cant_desc'],
                row['despacho__transportista__nombre_transportista'],
                empresa,  # mostrar el nombre comercial
                row['sacos_pendientes_de_pago'],
                row['total_a_pagar']
            ])

        # Fila final con total
        ultima_fila = ws.max_row + 1
        ws.cell(row=ultima_fila, column=10, value="TOTAL GENERAL:")
        ws.cell(row=ultima_fila, column=11, value=round(total_general, 2))

        # Guardar y responder
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"reporte_estiba_{fecha_inicio}_{fecha_fin}.xlsx"
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

def sanear_y_procesar_data(data):

    empresa_bd = data['ordenes_compra'][0].get('empresa').get('nombre_empresa')
    empresa = 'NO DETECTADA'
    if empresa_bd == "bd_trading_starsoft":
        empresa = "TRADING SEMILLA SAC"
    elif empresa_bd == "bd_semilla_starsoft":
        empresa = "LA SEMILLA DE ORO SAC"
    elif empresa_bd == "bd_maxi_starsoft":
        empresa = "MAXIMILIAN INVERSIONES SA"
    data['empresa'] = empresa

    config_despacho = data["configuracion_despacho"][0]
    # Variables
    suma_peso_salida_kg = 0.00
    suma_peso_llegada_kg = 0.00
    total_sacos_cargados = 0
    total_sacos_descargados = 0
    total_sacos_faltantes = 0
    total_sacos_rotos = 0
    total_sacos_humedos = 0
    total_sacos_mojados = 0
    merma_total = 0.00
    pago_estiba_list = []
    total_descuento_estiba = 0.00
    otros_gastos = []


    # Diccionario para contar las ocurrencias de cada tipo de pago_estiba
    conteo_pago_estiba = {
        "No pago estiba": 0,
        "Pago parcial": 0,
        "Pago estiba": 0,
        "Transbordo": 0
    }


    for item in data['detalle_despacho']:
        try:
            suma_peso_salida_kg += float(str(item["peso_salida"]).replace(",", ""))
            suma_peso_llegada_kg += float(str(item["peso_llegada"]).replace(",", ""))
            total_sacos_cargados += int(str(item["sacos_cargados"]))
            total_sacos_descargados += int(str(item["sacos_descargados"]))
            merma_total += float(str(item["merma"]).replace(",", ""))
        except (ValueError, AttributeError) as e:
            print(f"Error procesando item: {item}, error: {e}")
        total_sacos_faltantes += int(item["sacos_faltantes"])
        total_sacos_rotos += int(item["sacos_rotos"])  # Contar sacos rotos
        total_sacos_humedos += int(item["sacos_humedos"])  # Contar sacos húmedos
        total_sacos_mojados += int(item["sacos_mojados"])  # Contar sacos mojado

        # Evaluar el valor de pagoEstiba
        pago_estiba_value = item.get("pago_estiba")

        if pago_estiba_value is None:
            print(f"Advertencia: El campo 'pagoEstiba' está vacío para el item {item}")
            continue

        # Contar la cantidad de cada tipo de pago_estiba
        if pago_estiba_value in conteo_pago_estiba:
            conteo_pago_estiba[pago_estiba_value] += 1

        # Caso 1: "No pago 100 bolivianos"
        if "No pago estiba" in pago_estiba_value:
            # Concatenar placaLlegada y pagoEstiba
            # formatted_value = f"{item['placaLlegada']} - {pago_estiba_value}"
            # Agregar el resultado a la lista con un valor calculado
            pago_estiba_list.append({
                "placa": item['placa_llegada'],
                "detalle": pago_estiba_value,
                "monto_descuento": calcular_monto_descuento_estiba(item["sacos_descargados"],item["sacos_descargados"],config_despacho["tipo_cambio_desc_ext"])
            })

        # Caso 2: "Pago parcial"
        elif "Pago parcial" in pago_estiba_value:
            pago_estiba_list.append({
                "placa": item['placa_llegada'],
                "detalle": f"No pago x {item['cant_desc']}",
                "monto_descuento": calcular_monto_descuento_estiba(item["sacos_descargados"],item["cant_desc"],config_despacho["tipo_cambio_desc_ext"])
            })



    for item in pago_estiba_list:
        total_descuento_estiba += item['monto_descuento']

    datos_de_costo = calcular_costo_por_kg(data["flete_pactado"], config_despacho["precio_prod"],
                                           config_despacho["margen_financiero"], config_despacho["gastos_nacionalizacion"])

    diferencia_peso_kg = suma_peso_salida_kg - suma_peso_llegada_kg

    peso_sacos_faltantes = calcular_peso_no_considerado_por_sacos_faltante(total_sacos_faltantes)

    diferencia_peso_por_cobrar = calcular_diferencia_de_peso_por_cobrar_kg(diferencia_peso_kg,
                                                                           config_despacho.get("merma_permitida"),
                                                                           peso_sacos_faltantes)

    descuento_sacos_rotos = total_sacos_rotos * config_despacho["precio_sacos_rotos"]
    descuento_sacos_humedos = total_sacos_humedos * config_despacho["precio_sacos_humedos"]
    descuento_sacos_mojados = total_sacos_mojados * config_despacho["precio_sacos_mojados"]
    flete_base = data["flete_pactado"] * (data["peso_neto_crt"] / 1000)  # Convertir kg a toneladas
    precio_por_tonelada = config_despacho["precio_prod"] * 1000
    precio_bruto = precio_por_tonelada + data["flete_pactado"] + config_despacho["margen_financiero"] + config_despacho[
        "gastos_nacionalizacion"]
    igv = round(datos_de_costo.get("igv"),2)
    precio_bruto_final = precio_bruto + igv
    precio_por_tonelada_final = round(precio_bruto_final / 1.18, 2)
    precio_por_kg = round(precio_bruto_final / 1000, 4)
    total_descuento_sacos_faltantes = round(peso_sacos_faltantes * precio_por_kg, 2)
    total_descuento_sacos = calcular_descuento_sacos(total_descuento_sacos_faltantes, descuento_sacos_rotos,
                                                     descuento_sacos_humedos, descuento_sacos_mojados)
    total_descuento_solo_sacos = calcular_descuento_solo_sacos(descuento_sacos_rotos, descuento_sacos_humedos,
                                                               descuento_sacos_mojados)
    pago_final = flete_base - total_descuento_sacos


    # Paso 3: Calcular el descuento por diferencia de peso
    # Si la diferencia de peso por cobrar es positiva, calculamos el descuento
    if diferencia_peso_por_cobrar and diferencia_peso_por_cobrar > 0:
        descuento_por_diferencia_peso = diferencia_peso_por_cobrar * precio_por_kg
    else:
        descuento_por_diferencia_peso = 0  # No hay descuento si no hay diferencia de peso

    # Si hay diferencia de peso por cobrar, sumamos al pago final el monto correspondiente
    if diferencia_peso_por_cobrar and diferencia_peso_por_cobrar > 0:
        pago_final -= descuento_por_diferencia_peso

    # Si hay descuento por pago de estiba
    if total_descuento_estiba and total_descuento_estiba > 0:
        pago_final -= total_descuento_estiba

    # si hay descuento por sacos faltantes:
    if total_descuento_sacos and total_descuento_sacos > 0:
        pago_final -= total_descuento_sacos

    aux_descuento = descuento_por_diferencia_peso + total_descuento_sacos_faltantes + descuento_sacos_rotos + descuento_sacos_humedos + descuento_sacos_mojados

    total_otros_gastos=0.00
    for item in data['gastos_extra']:
        try:
            monto = float(item.get('monto', 0.00) or 0.00)
        except ValueError:
            monto = 0.00

        total_otros_gastos += monto
        otros_gastos.append({
            "id": item.get("id"),
            "descripcion": item.get("descripcion", ""),
            "monto": monto  # ahora es float
        })


    total_dsct_sin_gastos_otros = round((descuento_por_diferencia_peso + total_descuento_estiba + total_descuento_sacos_faltantes + descuento_sacos_rotos + descuento_sacos_humedos + descuento_sacos_mojados),2)

    total_dsct_con_otros_gastos = round((descuento_por_diferencia_peso +
                                         total_descuento_estiba + total_descuento_sacos_faltantes +
                                         descuento_sacos_rotos + descuento_sacos_humedos +
                                         descuento_sacos_mojados + total_otros_gastos),
                                        2)

    total_dsct = round((descuento_por_diferencia_peso + total_descuento_estiba + total_descuento_sacos_faltantes + descuento_sacos_rotos + descuento_sacos_humedos + descuento_sacos_mojados + total_otros_gastos),2)

    total_a_pagar = round(flete_base - total_dsct, 2)

    total_global_descuentos = round(total_descuento_sacos + total_descuento_estiba + descuento_por_diferencia_peso, 2)


    data['producto']=data['ordenes_compra'][0].get('producto').get('nombre_producto')
    data['total_sacos_cargados']=total_sacos_cargados
    data['total_sacos_descargados'] = total_sacos_descargados
    data['total_sacos_faltantes'] =total_sacos_faltantes
    data['total_sacos_rotos'] = total_sacos_rotos  # Contar sacos rotos
    data['total_sacos_humedos'] = total_sacos_humedos
    data['total_sacos_mojados'] = total_sacos_mojados
    data['suma_peso_salida_kg'] = suma_peso_salida_kg
    data['suma_peso_llegada_kg'] = suma_peso_llegada_kg
    data['merma_total'] = merma_total
    data['pago_estiba_list']=pago_estiba_list
    data['diferencia_de_peso']=diferencia_peso_kg
    data['peso_sacos_faltante']=peso_sacos_faltantes
    data['diferencia_peso_por_cobrar']=diferencia_peso_por_cobrar
    data['descuento_sacos_rotos']= descuento_sacos_rotos
    data['descuento_sacos_humedos']= descuento_sacos_humedos
    data['descuento_sacos_mojados']= descuento_sacos_mojados
    data['flete_base'] = flete_base
    data['precio_por_tonelada'] = precio_por_tonelada
    data['precio_bruto'] = precio_bruto
    data['igv'] = igv
    data['precio_bruto_final']=precio_bruto_final
    data['precio_por_tonelada_final'] = precio_por_tonelada_final
    data['precio_por_kg']=precio_por_kg
    data["total_luego_dsctos_sacos"]= calcular_monto_luego_dsctos_sacos(flete_base, total_descuento_sacos, descuento_por_diferencia_peso)
    data['descuento_sacos_faltantes']=total_descuento_sacos_faltantes
    data['total_descuento_sacos'] = total_descuento_sacos
    data['total_descuento_solo_sacos']=total_descuento_solo_sacos
    data['pago_final']=pago_final
    data['descuento_peso_sacos_faltantes']= peso_sacos_faltantes
    data['descuento_por_diferencia_peso']=descuento_por_diferencia_peso
    data['total_descuento_estiba']=total_descuento_estiba
    data['aux_descuento'] = aux_descuento
    data['total_dsct_sin_gastos_otros']=total_dsct_sin_gastos_otros
    data['total_dsct_con_otros_gastos']=total_dsct_con_otros_gastos
    data['total_dsct']=total_dsct
    data['otros_gastos']=otros_gastos
    data['total_otros_gastos'] = total_otros_gastos
    data['total_a_pagar']=total_a_pagar
    data['total_global_descuentos']=total_global_descuentos
    data['conteo_pago_estiba']=conteo_pago_estiba

    return data

def generar_reporte_pdf_con_data_bd(data):
    try:
        buffer = io.BytesIO()  # Crear un buffer en memoria

        # Crear el PDF
        c = canvas.Canvas(buffer, pagesize=landscape(A4))

        c.setTitle(" Reporte cálculo de flete")

        # Obtener las dimensiones de la página
        width, height = landscape(A4)

        # primera linea
        c.setFont("Helvetica", 10)
        current_y = height - 40
        c.drawString(40, current_y, f"{data['empresa']}")

        current_datetime = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        c.setFont("Helvetica", 7)
        c.drawString(width - 130, current_y, f"{data['fecha_de_creacion']}")

        # segunda linea
        c.setFont("Helvetica-Bold", 10)
        texto = f"{data['producto']}"
        text_width = c.stringWidth(texto, "Helvetica", 12)
        # Calcular la posición X para centrar el texto
        x_position = (width - text_width) / 2 + 20
        # Dibujar el texto centrado
        current_y -= 20
        c.drawString(x_position, current_y, texto)

        # tercera linea
        current_y -= 20
        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, current_y, f"CARTA PORTE: {data['carta_porte']}")

        # CUARTA LINEA DATOS DEL DESPACHO
        ordenes_recojo = data['ordenes_despacho']
        ordenes_string = " / ".join([f"{row['orden_compra']['numero_oc']}({row['numero_recojo']})" for row in ordenes_recojo])
        textos = [
            f"N° DUA: {data['dua']}",
            f"Fec. Num.: {data['fecha_numeracion']}",
            f"Factura N°: {data['num_factura']}",
            f"OC: {ordenes_string}",
            f"CANT. {data['total_sacos_cargados']} sacos",
            f" {data['peso_neto_crt']} kg"
        ]
        # Fuente y tamaño
        font_name = "Helvetica"
        font_size = 8
        c.setFont(font_name, font_size)
        # Espacio entre columnas
        margin = 30
        # Posición inicial en Y (mantener fija)
        current_y -= 20
        y_position = current_y
        # Posición inicial en X
        x_position = 40
        # Dibujar todos los textos en una línea tomando en cuenta su tamaño para darle cierto espaciado
        for texto in textos:
            # Calcular el ancho de cada texto
            text_width = c.stringWidth(texto, font_name, font_size)

            # Dibujar el texto
            c.drawString(x_position, y_position, texto)

            # Actualizar la posición X para el siguiente texto
            x_position += text_width + margin

            # Si el texto excede el tamaño de la página en el eje X, detener la creación del contenido
            if x_position > width:
                break

        # QUINTA LINEA DONDE INICIA LA TABLA:
        current_y -= 20
        # Datos iniciales de la tabla #1 (Titulos y subtitulos)
        data_table = [
            # Línea de títulos
            ['FEC. INGRE.', 'EMPRESA DE TRANSP.', 'CARGA', '', '', '', 'DESCARGA', '', '', '', 'DESCUENTOS', '', '', '',
             ''],
            ['', '', 'N°', 'Placa S.', 'Sacos C.', 'Peso S.', 'Placa L.', 'Sacos D.', 'Peso L.', 'Merma', 'S. Falt.',
             'S. Rotos', 'S. Humed.', 'S. Mojad.', 'Estibaje'],
        ]
        styles = getSampleStyleSheet()
        parrafo = Paragraph(f"<para align=center spaceb=3><b>{data['transportista'].get('nombre_transportista')}</b></para>",
                            styles["BodyText"])
        for i, row in enumerate(data['detalle_despacho']):
            if i == 0:  # Primera fila (datos especiales)
                data_table.append([
                    f"{data['fecha_numeracion']}",
                    parrafo,
                    str(i+1),
                    row['placa_salida'],
                    str(row['sacos_cargados']),
                    str(row['peso_salida']),
                    row['placa_llegada'],
                    str(row['sacos_descargados']),
                    str(row['peso_llegada']),
                    str(row['merma']),
                    str(row['sacos_faltantes']),
                    str(row['sacos_rotos']),
                    str(row['sacos_humedos']),
                    str(row['sacos_mojados']),
                    str(row['pago_estiba']),
                ])
            else:  # Para el resto de las filas
                data_table.append([
                    "",  # Cadena vacía
                    "",  # Cadena vacía
                    str(i+1),
                    row['placa_salida'],
                    str(row['sacos_cargados']),
                    str(row['peso_salida']),
                    row['placa_llegada'],
                    str(row['sacos_descargados']),
                    str(row['peso_llegada']),
                    str(row['merma']),
                    str(row['sacos_faltantes']),
                    str(row['sacos_rotos']),
                    str(row['sacos_humedos']),
                    str(row['sacos_mojados']),
                    str(row['pago_estiba']),
                ])
        # Agregamos fila de totales
        data_table.append(
            ['', f"Flete x TM: $ {data['flete_pactado']:.2f}", f"{''}", 'TOTAL',
             f"{data['total_sacos_cargados']}", f"{data['suma_peso_salida_kg']}", 'TOTAL',
             f"{data['total_sacos_descargados']}", f"{data['suma_peso_llegada_kg']}",
             f"{data['merma_total']}", f"{data['total_sacos_faltantes']}",
             f"{data['total_sacos_rotos']}", f"{data['total_sacos_humedos']}",
             f"{data['total_sacos_mojados']}"])
        # Ancho de columnas
        col_widths = [60, 110, 20, 60, 40, 60, 60, 60, 40, 60, 30, 30, 30, 30, 70]
        # Crear la tabla
        table = Table(data_table, colWidths=col_widths, rowHeights=15)
        color_rosa = Color(red=250 / 255, green=210 / 255, blue=202 / 255)
        color_celeste = Color(red=176 / 255, green=225 / 255, blue=250 / 255)
        numero_filas = len(data_table)
        # Estilos para la tabla
        table_style = TableStyle([
            ('SPAN', (0, 0), (0, 1)),  # Fusionar "FEC. INGRE."
            ('SPAN', (1, 0), (1, 1)),  # Fusionar "EMPRESA DE TRANSP."
            ('SPAN', (2, 0), (5, 0)),  # Fusionar "CARGA"
            ('SPAN', (6, 0), (9, 0)),  # Fusionar "DESCARGA"
            ('SPAN', (10, 0), (14, 0)),  # Fusionar "DESCUENTOS"
            ('SPAN', (0, 2), (0, numero_filas - 1)),  # Fusionar "fecha"
            ('SPAN', (1, 2), (1, numero_filas - 2)),  # Fusionar "nombre empresa transporte"
            ('FONTSIZE', (0, 0), (-1, -1), 6),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Centrar todo
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Fondo gris para los títulos
            ('BACKGROUND', (0, 1), (-1, 1), colors.lightgrey),  # Fondo gris claro para subtítulos
            ('BACKGROUND', (2, 0), (5, numero_filas), colors.yellowgreen),  # Color "CARGA"
            ('BACKGROUND', (6, 0), (9, numero_filas), color_celeste),  # Color "DESCARGA"
            ('BACKGROUND', (10, 0), (14, numero_filas), color_rosa),  # Fusionar "DESCUENTOS"
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Negrita en títulos
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),  # Negrita en subtítulos
            # ('FONTSIZE', (15, 0), (15, numero_filas - 1), 5),
        ])
        # Aplicar estilos a la tabla
        table.setStyle(table_style)
        # Calcular la altura total de la tabla
        table_height = len(data_table) * 15  # 20 es la altura de cada fila
        # Calcular la posición de la tabla en la página
        table.wrapOn(c, width, height)
        table.drawOn(c, 40, current_y - table_height)  # Ajustar la posición dependiendo de la cantidad de filas

        current_y = current_y - table_height - 20

        # Data para la tabla #2 (Detalles de pesos)
        second_data_table = [
            ["Faltante peso Sta. Cruz - Desaguadero:", f"{data['diferencia_de_peso']:.2f} Kg."],
            ["Merma permitida:", f"{data['configuracion_despacho'][0]['merma_permitida']:.2f} Kg."],  #POSIBLE ERROR
            ["Desc. diferencia de peso:", f"{data['diferencia_peso_por_cobrar']:.2f} Kg."],
            ["Desc. sacos falantes:", f"{data['descuento_peso_sacos_faltantes']:.2f} Kg."]
        ]
        second_col_widths = [110, 50]
        second_table = Table(second_data_table, colWidths=second_col_widths, rowHeights=14)
        second_table_style = TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 6),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Centrar
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
        ])
        second_table.setStyle(second_table_style)
        second_table_height = len(second_data_table) * 14
        second_table.wrapOn(c, width, height)
        second_table.drawOn(c, 40, current_y - second_table_height)

        # Data para la tabla #3 (Resumen de descuentos)
        third_data_table = [
            ["FLETE", f"$ {data['flete_base']:.2f}"],
            ["Dscto. por dif. de peso", f"$ {data['descuento_por_diferencia_peso']:.2f}"],
            ["Dscto. por sacos faltantes", f"$ {data['descuento_sacos_faltantes']:.2f}"],
            ["Dscto. por sacos rotos", f"$ {data['descuento_sacos_rotos']:.2f}"],
            ["Dscto. por sacos humedos", f"$ {data['descuento_sacos_humedos']:.2f}"],
            ["Dscto. por sacos mojados", f"$ {data['descuento_sacos_mojados']:.2f}"],
            ["FLETE TOTAL", f"$ {data['total_luego_dsctos_sacos']:.2f}"],
        ]
        # posicion en x para la tercera tabla:
        x_position_for_third_table = 260
        third_col_widths = [110, 40]
        third_table = Table(third_data_table, colWidths=third_col_widths, rowHeights=14)
        third_table_style = TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
        ])
        third_table.setStyle(third_table_style)
        third_table_height = len(third_data_table) * 14
        third_table.wrapOn(c, width, height)
        third_table.drawOn(c, x_position_for_third_table, current_y - third_table_height)

        # PARA VALIDACION DEL VALOR CRT
        peso_total_toneladas = data['peso_neto_crt'] / 1000
        flete_pactado = data["flete_pactado"]
        monto_flete = round((peso_total_toneladas * flete_pactado), 2)
        flete_base = data.get('flete_base',0.00)
        flete_base = round(flete_base, 2)

        estado_crt = ""
        if (flete_base <= monto_flete):
            estado_crt = "NO SOBREPASA VALOR CRT"
        else:
            estado_crt = "SOBREPASA VALOR CRT"
        c.setFont("Helvetica-Bold", 7)
        c.drawString(x_position_for_third_table + 150 + 10, current_y - 8, estado_crt)

        # Primera LLave para descuento sacos RMH
        c.setLineWidth(0.5)
        first_x_position_for_key = x_position_for_third_table + 150 + 10
        first_key_y_position = current_y - (14 * 3)
        last_key_y_position = first_key_y_position - (14 * 3)
        middle_of_key = (first_key_y_position + last_key_y_position) / 2
        inclinacion = 2
        c.line(first_x_position_for_key, first_key_y_position - inclinacion, first_x_position_for_key,
               last_key_y_position + inclinacion)
        c.line(first_x_position_for_key, first_key_y_position - inclinacion, first_x_position_for_key - inclinacion,
               first_key_y_position)
        c.line(first_x_position_for_key - inclinacion, last_key_y_position, first_x_position_for_key,
               last_key_y_position + inclinacion)
        c.line(first_x_position_for_key, middle_of_key, first_x_position_for_key + inclinacion, middle_of_key)
        c.setFont("Helvetica-Bold", 7)
        c.drawString(first_x_position_for_key + inclinacion + 4, middle_of_key,
                     f"$ {data['total_descuento_solo_sacos']:.2f}")

        # Segunda llave para descuento de sacos RMH
        c.setLineWidth(0.5)

        # Coordenadas para dibujar segunda la llave
        key_start_x = first_x_position_for_key + inclinacion + 4 + 40
        key_start_y = current_y - 14
        key_end_y = key_start_y - (14 * 5)
        key_middle_y = (key_start_y + key_end_y) / 2
        key_inclination = 2

        # Dibujar la llave
        c.line(key_start_x, key_start_y - key_inclination, key_start_x, key_end_y + key_inclination)
        c.line(key_start_x, key_start_y - key_inclination, key_start_x - key_inclination, key_start_y)
        c.line(key_start_x - key_inclination, key_end_y, key_start_x, key_end_y + key_inclination)
        c.line(key_start_x, key_middle_y, key_start_x + key_inclination, key_middle_y)

        # Texto del descuento
        c.setFont("Helvetica-Bold", 7)
        c.drawString(key_start_x + key_inclination + 4, key_middle_y,
                     f"$ {data['aux_descuento']:.2f}")

        # Lista de placas y estado de estiba:
        y_position_for_list = current_y - third_table_height - 10
        c.setFont("Helvetica", 6)
        for item in data["pago_estiba_list"]:
            # Crear texto en una línea
            c.drawString(x_position_for_third_table, y_position_for_list,
                         f"{item['placa']} {item['detalle']}")  # Placa y detalle
            c.drawString(x_position_for_third_table + 130, y_position_for_list,
                         f"$ {item['monto_descuento']:.2f}")  # Monto alineado a la derecha
            y_position_for_list -= 10  # Reducir la posición vertical para la próxima línea

        for item in data["otros_gastos"]:
            c.drawString(x_position_for_third_table, y_position_for_list, f"{item['descripcion']}")
            c.drawString(x_position_for_third_table + 130, y_position_for_list, f"$ {item['monto']:.2f}")
            y_position_for_list -= 10

        # Coordenadas para dibujar la tercera llave
        tirth_key_start_x = first_x_position_for_key
        tirth_key_start_y = current_y - third_table_height - 5
        tirth_key_end_y = y_position_for_list + 10
        tirth_key_middle_y = (tirth_key_start_y + tirth_key_end_y) / 2
        tirth_key_inclination = 2
        # Dibujar la tercera llave
        c.line(tirth_key_start_x, tirth_key_start_y - tirth_key_inclination, tirth_key_start_x,
               tirth_key_end_y + tirth_key_inclination)
        c.line(tirth_key_start_x, tirth_key_start_y - tirth_key_inclination, tirth_key_start_x - tirth_key_inclination,
               tirth_key_start_y)
        c.line(tirth_key_start_x - tirth_key_inclination, tirth_key_end_y, tirth_key_start_x,
               tirth_key_end_y + tirth_key_inclination)
        c.line(tirth_key_start_x, tirth_key_middle_y, tirth_key_start_x + tirth_key_inclination, tirth_key_middle_y)
        # Texto del descuento
        total_decuento = data['total_descuento_estiba'] + data['total_otros_gastos']
        c.setFont("Helvetica-Bold", 7)
        c.drawString(tirth_key_start_x + tirth_key_inclination + 4, tirth_key_middle_y - 2,
                     f"$ {total_decuento:.2f}")

        # Escribir linea de neto a pagar
        table_total_data = [
            ["TOTAL A PAGAR:", f"${data['total_a_pagar']:.2f}"]
        ]
        # Crear la tabla
        table_total = Table(table_total_data, colWidths=[110, 40])  # Anchos de columnas
        # Aplicar estilo a la tabla
        table_total.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.orange),  # Fondo naranja
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # Texto en color negro
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alineación izquierda primera columna
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),  # Fuente en negrita
            ('FONTSIZE', (0, 0), (-1, -1), 8),  # Tamaño de la fuente
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alineación derecha segunda columna
        ]))
        altura_tabla_total = len(table_total_data) * 14
        table_total.wrapOn(c, width, height)
        table_total.drawOn(c, x_position_for_third_table, y_position_for_list - altura_tabla_total - 20)

        # Data para la tabla #4 (resumen de descuentos)
        new_x_position = 260 + 150 + 150
        fourth_data_table = [
            ["FLETE", f"$ {data['flete_base']:.2f}"],
            ["Dscto.", f"$ {data['total_dsct']:.2f}"],
            ["Neto.", f"$ {data['total_a_pagar']:.2f}"],
        ]
        fourth_col_widths = [30, 50]
        fourth_table = Table(fourth_data_table, colWidths=fourth_col_widths, rowHeights=14)
        fourth_table_style = TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
        ])
        fourth_table.setStyle(fourth_table_style)
        fourth_table_height = len(fourth_data_table) * 14
        fourth_table.wrapOn(c, width, height)
        fourth_table.drawOn(c, new_x_position, current_y - fourth_table_height - (14 * 3))
        c.setFont("Helvetica-Bold", 7)
        c.drawString(new_x_position, current_y - fourth_table_height + 5, "RESUMEN")

        # Data para la tabla #5 (determinacion de costos)
        second_new_x_position = new_x_position + 120
        c.setFont("Helvetica-Bold", 6)
        c.drawString(second_new_x_position, current_y + 5, "Determinacion de costo x Kg.")
        fifth_data_table = [
            ["C.P", f"{data['precio_por_tonelada']:.2f}"],
            ["F.B", f"$ {data['flete_pactado']:.2f}"],
            ["G.N", f"{data['configuracion_despacho'][0]['gastos_nacionalizacion']}"],
            ["MF CIA", f"{data['configuracion_despacho'][0]['margen_financiero']}"],
            ["IGV 18%", f"{data['igv']}"],
            ["PRECIO DES", f"{data['precio_bruto_final']:.2f}"],
            ["COSTO TM", f"{data['precio_por_tonelada_final']}"],
            ["COSTO Kg", f"{data['precio_por_kg']}"],
        ]
        fifth_col_widths = [50, 40]
        fifth_table = Table(fifth_data_table, colWidths=fifth_col_widths, rowHeights=14)
        fifth_table_style = TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir cuadrícula
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear la segunda columna a la derecha
        ])
        fifth_table.setStyle(fifth_table_style)
        fifth_table_height = len(fifth_data_table) * 14
        fifth_table.wrapOn(c, width, height)
        fifth_table.drawOn(c, second_new_x_position, current_y - fifth_table_height)

        # Espacio de separación entre tablas
        separation_space = 20  # Espacio en píxeles

        # Posición Y de la nueva tabla (debajo de la anterior)
        new_table_y_position = current_y - fifth_table_height - separation_space

        # Data para la nueva tabla (Conteo de 'pago_estiba')
        c.setFont("Helvetica-Bold", 6)
        c.drawString(second_new_x_position, new_table_y_position + 5, "Resumen de estibaje")

        conteo_pago_estiba=data["conteo_pago_estiba"]
        payment_summary_table = [
            ["No pago estiba", conteo_pago_estiba["No pago estiba"]],
            ["Pago parcial", conteo_pago_estiba["Pago parcial"]],
            ["Pago estiba", conteo_pago_estiba["Pago estiba"]],
            ["Transbordo", conteo_pago_estiba["Transbordo"]],
        ]

        payment_col_widths = [70, 40]
        payment_table = Table(payment_summary_table, colWidths=payment_col_widths, rowHeights=14)

        payment_table_style = TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 6),  # Tamaño de fuente
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alinear toda la tabla a la izquierda
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Cuadrícula
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Alinear los valores a la derecha
        ])

        payment_table.setStyle(payment_table_style)

        # Calcular altura de la tabla de resumen
        payment_table_height = len(payment_summary_table) * 14

        # Dibujar la nueva tabla en el PDF
        payment_table.wrapOn(c, width, height)
        payment_table.drawOn(c, second_new_x_position, new_table_y_position - payment_table_height)


        # SALTO DE LINEA PARA ESPACIO DE DESCUENTOS
        current_y = current_y - second_table_height - 20

        x_start = 40  # Eje X constante
        line_height = 10  # Interlineado
        # Escribir título
        c.setFont("Helvetica-Bold", 9)
        c.drawString(x_start, current_y, "Descuentos:")
        current_y -= line_height + 5  # Reducir Y para el siguiente contenido
        # Cambiar la fuente para los detalles
        c.setFont("Helvetica", 7)
        # Condicionales y datos
        if data["descuento_por_diferencia_peso"] > 0:
            c.drawString(
                x_start,
                current_y,
                f"B/V 004-:        dscto. x dif. peso $ {data['descuento_por_diferencia_peso']:.2f}"
            )
            current_y -= line_height

        if data["descuento_sacos_faltantes"] > 0:
            c.drawString(
                x_start,
                current_y,
                f"B/V 004-:        dscto. x sacos falt. $ {data['descuento_sacos_faltantes']:.2f}"
            )
            current_y -= line_height

        if data["total_descuento_solo_sacos"] > 0:
            c.drawString(
                x_start,
                current_y,
                f"B/V 004-:        dscto. x sacos R,H,M. $ {data['total_descuento_solo_sacos']:.2f}"
            )
            current_y -= line_height

        if data["total_descuento_estiba"] > 0:
            c.drawString(
                x_start,
                current_y,
                f"B/V 004-:        dscto. x estibaje $ {data['total_descuento_estiba']:.2f}"
            )
            current_y -= line_height

        for gasto in data['otros_gastos']:
            descripcion = gasto.get("descripcion", "")
            monto = gasto.get("monto", 0.00)

            c.drawString(
                x_start,
                current_y,
                f"B/V 004-:        {descripcion} $ {monto:.2f}"
            )
            current_y -= line_height



        # Total
        c.setFont("Helvetica-Bold", 8)
        c.drawString(
            x_start,
            current_y - 5,
            f"TOTAL A DESCONTAR:   $ {data['total_dsct_con_otros_gastos']:.2f}"
        )

        # Guardar el PDF en el buffer
        c.showPage()
        c.save()

        buffer.seek(0)  # Volver al inicio del buffer

        return buffer.getvalue()  # Devolver el PDF en binario

    except KeyError as e:
        print(f"Error desde gen_reporte: Clave faltante en los datos - {e}")
        return None  # O podrías lanzar la excepción de nuevo con: raise
    except Exception as e:
        print(f"Error inesperado: {e}")
        return None  # O manejarlo según tus necesidades

class DespachoDeleteView(generics.DestroyAPIView):
    """
    API endpoint para eliminar despachos.
    Requiere permisos de administrador de importaciones.
    """
    queryset = Despacho.objects.all()
    serializer_class = DespachoSerializer
    permission_classes = [IsAuthenticated, CanAccessImportaciones, CanDeleteResource]

    def delete(self, request, pk):
        try:
            despacho = get_object_or_404(Despacho, pk=pk)

            # Serializamos los datos antes de eliminar
            data_despacho = DespachoSerializer(despacho).data

            # Eliminamos el despacho
            despacho.delete()

            return JsonResponse(
                {
                    "success": True,
                    "message": f"Despacho {pk} eliminado correctamente",
                    "deleted_data": data_despacho
                },
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return JsonResponse(
                {
                    "success": False,
                    "error": str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ProcesarArchivoView(APIView):
    """
    API endpoint para procesar archivos de documentos.
    Requiere permisos para editar documentos.
    """
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAuthenticated, CanEditDocuments]

    def post(self, request, *args, **kwargs):
        archivo = request.FILES.get('archivo')
        numero_dam = request.data.get('numero_dam')  # Recibir N° de DAM del frontend

        if not archivo:
            return Response({'error': 'No se ha enviado ningún archivo'}, status=400)

        # Guardar el archivo temporalmente
        ruta_archivo = os.path.join(settings.MEDIA_ROOT, archivo.name)
        with open(ruta_archivo, 'wb+') as destino:
            for chunk in archivo.chunks():
                destino.write(chunk)

        # Procesar archivo RAR
        try:
            with rarfile.RarFile(ruta_archivo) as rf:
                file_list = rf.namelist()

            # Obtener solo carpetas
            folder_list = sorted(set(os.path.dirname(f) for f in file_list if f.endswith("/")))
            num_folders = len(folder_list)

            # Si el N° de DAM está presente, renombrar la carpeta principal
            if numero_dam:
                nueva_carpeta = f"DAM_{numero_dam}"
                destino_final = os.path.join(settings.MEDIA_ROOT, nueva_carpeta)

                if not os.path.exists(destino_final):
                    os.makedirs(destino_final)

                os.rename(ruta_archivo, os.path.join(destino_final, archivo.name))
                return Response({
                    'mensaje': 'Archivo procesado y renombrado correctamente',
                    'cantidad_carpetas': num_folders,
                    'carpetas': folder_list,
                    'ruta_guardada': destino_final
                })

            return Response({
                'mensaje': 'Archivo procesado correctamente',
                'cantidad_carpetas': num_folders,
                'carpetas': folder_list
            })

        except rarfile.BadRarFile:
            return Response({'error': 'El archivo no es un RAR válido'}, status=400)

class GuardarArchivoView(APIView):
    """
    API endpoint para guardar archivos.
    Requiere permisos para editar documentos.
    """
    permission_classes = [IsAuthenticated, CanEditDocuments]
    def post(self, request, *args, **kwargs):
        archivos = request.data.get("archivos", [])  # Lista de archivos renombrados

        if not archivos:
            return Response({'error': 'No se han proporcionado archivos para guardar'}, status=400)

        base_path = settings.MEDIA_ROOT  # Directorio base donde se guardarán los archivos

        for archivo in archivos:
            original = archivo.get("original")
            nuevo = archivo.get("nuevo")

            if not original or not nuevo:
                continue

            ruta_original = os.path.join(base_path, original)
            ruta_nueva = os.path.join(base_path, nuevo)

            if os.path.exists(ruta_original):
                os.rename(ruta_original, ruta_nueva)

        return Response({'mensaje': 'Archivos guardados correctamente'})

class CargaDirectaView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        numero = request.data.get("numero_dua")
        anio = request.data.get("anio")
        archivos = request.FILES.getlist("archivos")

        if not numero or not anio:
            return Response({"error": "Debe proporcionar número y año de la DUA"}, status=400)

        if not archivos:
            return Response({"error": "No se proporcionaron archivos"}, status=400)

        # 🔧 Normalizar número antes de buscar
        numero_normalizado = numero.lstrip("0")

        try:
            declaracion, _ = Declaracion.objects.get_or_create(numero=numero_normalizado, anio=anio)
        except IntegrityError:
            # Si ocurre una condición de carrera, intenta obtenerlo directamente
            declaracion = Declaracion.objects.get(numero=numero_normalizado, anio=anio)

        repetidos = []

        for archivo in archivos:
            hash_archivo = calcular_hash_archivo(archivo)

            documento_existente = Documento.objects.filter(hash_archivo=hash_archivo).first()
            if documento_existente:
                declaracion_existente = documento_existente.content_object
                repetidos.append({
                    "archivo": archivo.name,
                    "registrado_en": f"{declaracion_existente.numero}-{declaracion_existente.anio}" if isinstance(declaracion_existente, Declaracion) else "Desconocido"
                })
                continue

            documento = Documento(
                content_object=declaracion,
                nombre_original=archivo.name,
                hash_archivo=hash_archivo,
                usuario=request.user
            )
            documento.archivo.save(archivo.name, archivo)
            documento.save()

        if repetidos:
            return Response({
                "mensaje": "Carga completada con advertencias",
                "archivos_omitidos": repetidos
            })
        return Response({"mensaje": "Archivos cargados correctamente"})

class ProcesarArchivoComprimidoView(APIView):
    """
    API endpoint para procesar archivos comprimidos.
    Requiere permisos para editar documentos.
    """
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAuthenticated, CanEditDocuments]

    def post(self, request):
        archivo = request.FILES.get("archivo")

        if not archivo:
            return Response({"error": "No se proporcionó archivo comprimido"}, status=400)

        os.makedirs(settings.MEDIA_ROOT, exist_ok=True)
        temp_path = os.path.join(settings.MEDIA_ROOT, archivo.name)
        with open(temp_path, 'wb+') as f:
            for chunk in archivo.chunks():
                f.write(chunk)

        try:
            if zipfile.is_zipfile(temp_path):
                with zipfile.ZipFile(temp_path) as zf:
                    folder_names = sorted(set(os.path.dirname(f) for f in zf.namelist() if f.endswith("/")))
            elif rarfile.is_rarfile(temp_path):
                with rarfile.RarFile(temp_path) as rf:
                    folder_names = sorted(set(os.path.dirname(f) for f in rf.namelist() if f.endswith("/")))
            else:
                return Response({"error": "Archivo no es ZIP ni RAR válido"}, status=400)

            return Response({
                "mensaje": "Archivo procesado",
                "carpetas": folder_names,
                "archivo_temp": archivo.name
            })

        except Exception as e:
            return Response({"error": str(e)}, status=500)

class AsignarDeclaracionDesdeComprimidoView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        archivo_temp = request.data.get("archivo_temp")
        asignaciones = request.data.get("asignaciones", [])

        if not archivo_temp or not asignaciones:
            return Response({"error": "Datos incompletos"}, status=400)

        temp_path = os.path.join(settings.MEDIA_ROOT, archivo_temp)

        if not os.path.exists(temp_path):
            return Response({"error": "Archivo temporal no encontrado"}, status=404)

        try:
            if zipfile.is_zipfile(temp_path):
                archivo_comp = zipfile.ZipFile(temp_path)
            elif rarfile.is_rarfile(temp_path):
                archivo_comp = rarfile.RarFile(temp_path)
            else:
                return Response({"error": "Archivo no válido"}, status=400)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

        repetidos = []

        for asignacion in asignaciones:
            carpeta = asignacion.get("carpeta")
            numero = asignacion.get("numero_dua")
            anio = asignacion.get("anio")
            numero = numero.lstrip('0')

            if not carpeta or not numero or not anio:
                continue

            try:
                numero = numero.lstrip('0')  # <-- normalizar antes de buscar
                declaracion, created = Declaracion.objects.get_or_create(numero=numero, anio=anio)
            except Declaracion.MultipleObjectsReturned:
                declaracion = Declaracion.objects.filter(numero=numero, anio=anio).first()
            except Declaracion.DoesNotExist:
                return Response(
                    {"error": f"No se pudo encontrar ni crear la declaración {numero}-{anio}."},
                    status=404
                )

            for file_name in archivo_comp.namelist():
                if file_name.startswith(carpeta) and not file_name.endswith("/"):
                    contenido = archivo_comp.read(file_name)
                    nombre_simple = os.path.basename(file_name)

                    hash_archivo = hashlib.sha256(contenido).hexdigest()
                    existente = Documento.objects.filter(hash_archivo=hash_archivo).first()

                    if existente:
                        declaracion_existente = existente.content_object  # Cambiado a content_object
                        repetidos.append({
                            "archivo": nombre_simple,
                            "registrado_en": f"{declaracion_existente.numero}-{declaracion_existente.anio}" if isinstance(declaracion_existente, Declaracion) else "Desconocido"
                        })
                        continue

                    carpeta_destino = os.path.join(settings.MEDIA_ROOT, "documentos", f"{numero}-{anio}")
                    os.makedirs(carpeta_destino, exist_ok=True)
                    ruta_destino = os.path.join(carpeta_destino, nombre_simple)

                    with open(ruta_destino, "wb") as f:
                        f.write(contenido)

                    file_content = ContentFile(contenido)

                    documento = Documento(
                        content_object=declaracion,
                        nombre_original=nombre_simple,
                        hash_archivo=hash_archivo,
                        usuario=request.user
                    )
                    # Esto activa la lógica de `upload_to=ruta_documento`
                    documento.archivo.save(nombre_simple, file_content)
                    documento.save()

        archivo_comp.close()
        os.remove(temp_path)

        if repetidos:
            return Response({
                "mensaje": "Carga completada con advertencias",
                "archivos_omitidos": repetidos
            })

        return Response({"mensaje": "Archivos extraídos y asignados correctamente"})

class ListarDeclaracionesView(APIView):
    """
    API endpoint para listar declaraciones.
    Requiere acceso al módulo de importaciones.
    """
    permission_classes = [IsAuthenticated, CanAccessImportaciones]
    
    def get(self, request):
        numero = request.query_params.get('numero')
        queryset = Declaracion.objects.prefetch_related('documentos').all()
        if numero:
            queryset = queryset.filter(numero__icontains=numero)
        data = DeclaracionConDocumentosSerializer(queryset, many=True).data
        return Response(data)

class DescargarZipView(APIView):
    """
    API endpoint para descargar archivos ZIP de declaraciones.
    Requiere acceso al módulo de importaciones.
    """
    permission_classes = [IsAuthenticated, CanAccessImportaciones]
    
    def get(self, request, numero, anio):
        try:
            declaracion = Declaracion.objects.get(numero=numero, anio=anio)
        except Declaracion.DoesNotExist:
            raise Http404("Declaración no encontrada")

        carpeta = f"{numero}-{anio}"

        # Obtener el content_type para el modelo Declaracion
        content_type = ContentType.objects.get_for_model(Declaracion)

        # Filtrar documentos usando la relación genérica
        documentos = Documento.objects.filter(
            content_type=content_type,
            object_id=declaracion.id
        )

        # Usar un buffer en memoria
        buffer = io.BytesIO()

        # Crear el archivo ZIP en memoria
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for doc in documentos:
                if default_storage.exists(doc.archivo.name):
                    nombre = os.path.basename(doc.archivo.name)
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.archivo.name)

                    with open(file_path, "rb") as f:
                        zip_file.writestr(nombre, f.read())

        buffer.seek(0)

        # Devolver el archivo ZIP como respuesta
        response = FileResponse(buffer, as_attachment=True, filename=f"{carpeta}.zip")
        return response

class EliminarDocumentoView(APIView):
    """
    API endpoint para eliminar documentos.
    Requiere permisos de eliminación y edición de documentos.
    """
    permission_classes = [IsAuthenticated, CanEditDocuments, CanDeleteResource]
    
    def delete(self, request, pk):
        try:
            doc = Documento.objects.get(pk=pk)
        except Documento.DoesNotExist:
            return Response(status=404)
        doc.archivo.delete(save=False)
        doc.delete()
        return Response(status=204)

class DocumentosPorDeclaracionView(APIView):
    """
    API endpoint para ver documentos por declaración.
    Requiere acceso al módulo de importaciones.
    """
    permission_classes = [IsAuthenticated, CanAccessImportaciones]
    def get(self, request, numero, anio):
        declaracion = get_object_or_404(Declaracion, numero=numero, anio=anio)

        # Obtener ContentType para el modelo Declaracion
        content_type = ContentType.objects.get_for_model(Declaracion)

        # Filtrar documentos relacionados a esa declaración
        documentos = Documento.objects.filter(
            content_type=content_type,
            object_id=declaracion.id
        )

        serializer = DocumentoSerializer(documentos, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class DocumentosPorDeclaracionUsuarioLogeadoView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, numero, anio):
        # 1. Obtener la declaración específica
        declaracion = get_object_or_404(Declaracion, numero=numero, anio=anio)

        # 2. Obtener ContentType del modelo Declaracion
        content_type = ContentType.objects.get_for_model(Declaracion)

        # 3. Filtrar los documentos relacionados a esa declaración y al usuario logeado
        documentos = Documento.objects.filter(
            content_type=content_type,
            object_id=declaracion.id,
            usuario=request.user
        )

        # 4. Serializar y devolver
        serializer = DocumentoSerializer(documentos, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class DescargarDocumentoView(APIView):
    """
    API endpoint para descargar documentos individuales.
    Requiere acceso al módulo de importaciones.
    """
    permission_classes = [IsAuthenticated, CanAccessImportaciones]
    def get(self, request, documento_id):
        documento = get_object_or_404(Documento, id=documento_id)
        file_path = documento.archivo.path
        file_name = documento.nombre_original or os.path.basename(file_path)

        mime_type, _ = mimetypes.guess_type(file_path)
        mime_type = mime_type or 'application/octet-stream'

        response = FileResponse(open(file_path, 'rb'), content_type=mime_type)
        response['Content-Disposition'] = f'attachment; filename="{smart_str(file_name)}"'
        return response

class ListarDeclaracionesDelUsuarioView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        numero = request.query_params.get('numero')

        # Filtra las declaraciones que tengan al menos un documento subido por el usuario autenticado
        queryset = Declaracion.objects.filter(
            documentos__usuario=request.user
        ).prefetch_related('documentos').distinct()

        # Filtro adicional por número si se proporciona
        if numero:
            queryset = queryset.filter(numero__icontains=numero)

        # Serializa y responde
        data = DeclaracionConDocumentosSerializer(queryset, many=True).data
        return Response(data)

class DocumentoVisualizarView(APIView):
    permission_classes = [IsAuthenticated] # Requiere que el usuario esté autenticado

    def get(self, request, pk, *args, **kwargs):
        """
        Permite a un usuario autenticado visualizar un documento específico.
        Se verifica que el documento exista y que el usuario tenga permisos
        para acceder a él.
        """
        try:
            documento = get_object_or_404(Documento, pk=pk)
        except Http404:
            return Response(
                {"detail": "Documento no encontrado."},
                status=status.HTTP_404_NOT_FOUND
            )

        # --- Lógica de Autorización (MUY IMPORTANTE) ---
        grupos_permitidos = ['importaciones_admin', 'importaciones_asis', 'proveedor']

        # 2. El usuario pertenece a un grupo específico (ej. "Administradores", "Gestores"):
        if not request.user.groups.filter(name__in=grupos_permitidos).exists():
            return Response({"detail": "No tienes permiso para visualizar este documento."},status=status.HTTP_403_FORBIDDEN)
        # 3. El usuario tiene un permiso específico (ej. 'tu_app_documentos.view_documento'):
        # if not request.user.has_perm('tu_app_documentos.view_documento', documento):
        #    return Response(
        #        {"detail": "No tienes permiso para visualizar este documento."},
        #        status=status.HTTP_403_FORBIDDEN
        #    )
        # 4. El documento está asociado a una Declaracion y el usuario tiene acceso a esa Declaracion:
        # if not request.user == documento.declaracion.usuario_propietario: # Asumiendo un campo 'usuario_propietario' en Declaracion
        #     return Response(
        #         {"detail": "No tienes permiso para visualizar documentos de esta declaración."},
        #         status=status.HTTP_403_FORBIDDEN
        #     )

        # Usaremos el ejemplo 1, que es el más directo con tu modelo actual,
        # asumiendo que 'usuario' en Documento es el que lo subió y tiene derecho a verlo.
        # Ajusta esta lógica según tus reglas de negocio.
        # -----------------------------------------------

        # Asegúrate de que el archivo exista en el sistema de archivos
        file_path = documento.archivo.path
        if not os.path.exists(file_path):
            return Response(
                {"detail": "El archivo no existe en el servidor."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Verifica que el archivo sea un PDF
        if not documento.nombre_original.lower().endswith('.pdf'):
            return Response(
                {"detail": "Solo se pueden visualizar archivos PDF."},
                status=status.HTTP_400_BAD_REQUEST # O 403, dependiendo de la política
            )

        # Abre el archivo y lo sirve como respuesta
        try:
            # Abre el archivo en modo binario
            # Usar 'inline' en Content-Disposition para que el navegador intente visualizarlo
            # Usar 'application/pdf' como Content-Type
            return FileResponse(
                open(file_path, 'rb'),
                content_type='application/pdf',
                as_attachment=False, # Esto es clave para 'inline' visualization
                filename=documento.nombre_original # Opcional, nombre sugerido para guardar
            )
        except Exception as e:
            return Response(
                {"detail": f"Error al servir el archivo: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class EditarPDFView(APIView):
    """
    API endpoint para editar archivos PDF.
    Requiere permisos para editar documentos.
    """
    permission_classes = [IsAuthenticated, CanEditDocuments]
    def post(self, request, pk):
        try:
            documento = Documento.objects.select_related('declaracion').get(pk=pk)
        except Documento.DoesNotExist:
            return Response({'error': 'Documento no encontrado'}, status=404)

        paginas_a_eliminar = request.data.get("paginas_a_eliminar", [])
        if not isinstance(paginas_a_eliminar, list):
            return Response({'error': 'Formato inválido de páginas'}, status=400)

        original_path = documento.archivo.path
        doc = fitz.open(original_path)

        if len(paginas_a_eliminar) >= doc.page_count:
            return Response({'error': 'No se puede eliminar todas las páginas'}, status=400)

        # Eliminar las páginas seleccionadas
        for pagina in sorted(set(paginas_a_eliminar), reverse=True):
            if 1 <= pagina <= len(doc):
                doc.delete_page(pagina - 1)

        # Generar nuevo nombre y ruta del archivo
        base_nombre, _ = os.path.splitext(documento.nombre_original)
        nuevo_nombre = f"{base_nombre}_editado.pdf"
        carpeta = f"{documento.declaracion.numero}-{documento.declaracion.anio}"
        nuevo_rel_path = os.path.join("documentos", carpeta, nuevo_nombre)
        nuevo_path = os.path.join(settings.MEDIA_ROOT, nuevo_rel_path)

        os.makedirs(os.path.dirname(nuevo_path), exist_ok=True)
        doc.save(nuevo_path)

        # Crear nuevo Documento
        nuevo_documento = Documento.objects.create(
            declaracion=documento.declaracion,
            archivo=nuevo_rel_path,
            nombre_original=nuevo_nombre,
            usuario=request.user if request.user.is_authenticated else None
        )

        # Crear registro en ExpedienteDeclaracion
        expediente = ExpedienteDeclaracion.objects.create(
            declaracion=documento.declaracion,
            documento=nuevo_documento,
            descripcion=f"Versión editada del documento ID {documento.id}. Páginas eliminadas: {paginas_a_eliminar}",
            tipo='edicion',
            usuario=request.user if request.user.is_authenticated else None
        )

        return Response({
            'mensaje': 'PDF editado con éxito',
            'nuevo_documento_id': nuevo_documento.id,
            'expediente_id': expediente.id,
            'nombre': nuevo_documento.nombre_original
        }, status=201)

@api_view(['GET'])
def obtener_pdf(request, pk):
    documento = get_object_or_404(Documento, pk=pk)
    return FileResponse(documento.archivo, content_type='application/pdf')

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pdf_list_from_server(request, documento_id):
    try:
        documento_actual = Documento.objects.get(id=documento_id)
    except Documento.DoesNotExist:
        return Response({"error": "Documento no encontrado"}, status=status.HTTP_404_NOT_FOUND)

    declaracion = documento_actual.declaracion

    documentos = Documento.objects.filter(declaracion=declaracion).exclude(id=documento_id)

    serializer = DocumentoListadoSerializer(documentos, many=True)
    return Response(serializer.data)

class DocumentosRelacionadosAPIView(APIView):
    """
    API endpoint para ver documentos relacionados.
    Requiere acceso al módulo de importaciones.
    """
    permission_classes = [IsAuthenticated, CanAccessImportaciones]
    def get(self, request, documento_id):
        try:
            # Obtiene el documento pasado por id
            documento = Documento.objects.get(id=documento_id)
        except Documento.DoesNotExist:
            return Response({"detail": "Documento no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        # Obtiene todos los documentos que tengan la misma declaracion
        documentos_relacionados = Documento.objects.filter(
            declaracion=documento.declaracion
        ).exclude(id=documento_id)

        serializer = DocumentoSerializer(documentos_relacionados, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class CombinarPDFsDeclaracionView(APIView):
    def post(self, request, numero, anio):
        return Response({"detail": "ok."}, status=status.HTTP_200_OK)

class AgregarDocumentosExistentesAPIView(APIView):
    def post(self, request, expediente_id):
        return Response({"detail": "ok"}, status=status.HTTP_200_OK)

class ReordenarPaginasAPIView(APIView):
    def post(self, request, expediente_id):
        nuevo_orden = request.data.get("nuevo_orden", [])  # Ej: [2, 0, 1] (índices desde 0)
        expediente = ExpedienteDeclaracion.objects.get(id=expediente_id)
        reader = PdfReader(expediente.archivo_pdf.path)
        writer = PdfWriter()

        for page_index in nuevo_orden:
            writer.add_page(reader.pages[page_index])

        output = io.BytesIO()
        writer.write(output)

        expediente.archivo_pdf.save(f"expediente_{expediente.id}_reordenado.pdf", ContentFile(output.getvalue()))
        expediente.save()

        return Response({"detail": "Páginas reordenadas correctamente."}, status=status.HTTP_200_OK)

class AsignarPaginasAPIView(APIView):
    def post(self, request):
        serializer = AsignarPaginasSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        documento_id = serializer.validated_data["documento_id"]
        asignaciones = serializer.validated_data["asignaciones"]

        try:
            documento = Documento.objects.get(id=documento_id)
        except Documento.DoesNotExist:
            return Response({"detail": "Documento no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        pdf_path = documento.archivo.path
        pdf_reader = PdfReader(pdf_path)
        paginas_usadas = set()

        for asignacion in asignaciones:
            page_num = asignacion["page"] - 1  # PyPDF2 usa índice base 0
            tipo_id = asignacion["tipo"]

            if page_num in paginas_usadas or page_num < 0 or page_num >= len(pdf_reader.pages):
                continue

            paginas_usadas.add(page_num)

            # Extraer la página y crear nuevo PDF
            pdf_writer = PdfWriter()
            pdf_writer.add_page(pdf_reader.pages[page_num])
            buffer = io.BytesIO()
            pdf_writer.write(buffer)
            buffer.seek(0)

            file_content = ContentFile(buffer.read())

            # Obtener nombre base sin extensión
            nombre_base = Path(documento.nombre_original).stem
            nuevo_nombre = f"{nombre_base}_p{page_num + 1}.pdf"

            # Validar existencia del tipo de documento
            try:
                tipo_documento = TipoDocumento.objects.get(id=tipo_id)
            except TipoDocumento.DoesNotExist:
                return Response({"detail": f"TipoDocumento con ID {tipo_id} no existe."}, status=400)

            # Crear expediente sin documento aún
            expediente = ExpedienteDeclaracion.objects.create(
                declaracion=documento.content_object,
                tipo=tipo_documento,
                usuario=request.user
            )

            # Crear y guardar el nuevo documento
            nuevo_documento = Documento(
                nombre_original=nuevo_nombre,
                usuario=request.user,
                content_object=expediente
            )
            nuevo_documento.archivo.save(nuevo_nombre, file_content)
            nuevo_documento.save()

            # Asociar el documento al expediente
            expediente.documento = nuevo_documento
            expediente.save()

        return Response({"detail": "Asignación de páginas completada."}, status=status.HTTP_201_CREATED)

class ListarExpedientesDeclaracionView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        numero = request.query_params.get('numero')
        anio_fiscal = request.query_params.get('anio_fiscal')
        mes_fiscal = request.query_params.get('mes_fiscal')

        queryset = ExpedienteDeclaracion.objects.select_related('declaracion')

        if numero:
            queryset = queryset.filter(declaracion__numero__icontains=numero.lstrip('0'))

        if anio_fiscal:
            queryset = queryset.filter(anio_fiscal=anio_fiscal)

        if mes_fiscal:
            queryset = queryset.filter(mes_fiscal=mes_fiscal)

        serializer = ExpedienteDeclaracionListadoSerializer(queryset, many=True)
        return Response(serializer.data)

class ListarExpedientesAgrupadosView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        numero = request.query_params.get('numero')
        anio_fiscal = request.query_params.get('anio_fiscal')
        mes_fiscal = request.query_params.get('mes_fiscal')
        empresa = request.query_params.get('empresa')

        queryset = ExpedienteDeclaracion.objects.select_related('declaracion')

        # Filtros
        if numero:
            queryset = queryset.filter(declaracion__numero__icontains=numero.lstrip("0"))
        if anio_fiscal:
            queryset = queryset.filter(anio_fiscal=anio_fiscal)
        if mes_fiscal:
            queryset = queryset.filter(mes_fiscal=mes_fiscal)
        if empresa:
            queryset = queryset.filter(empresa=empresa)

        agrupados = (
            queryset
            .values(
                'declaracion__id',
                'declaracion__numero',
                'declaracion__anio',
            )
            .annotate(
                cantidad_documentos=Count('id'),
                anio_fiscal=Min('anio_fiscal'),
                mes_fiscal=Min('mes_fiscal'),
                empresa=Min('empresa')  # Nueva línea
            )
            .order_by('-declaracion__anio', '-declaracion__numero')
        )

        resultado = [
            {
                'id': item['declaracion__id'],
                'numero_declaracion': item['declaracion__numero'],
                'anio_declaracion': item['declaracion__anio'],
                'cantidad_documentos': item['cantidad_documentos'],
                'anio_fiscal': item['anio_fiscal'],
                'mes_fiscal': item['mes_fiscal'],
                'empresa': item['empresa'],
            }
            for item in agrupados
        ]

        return Response(resultado)

class ListarDocumentosPorTipoView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, declaracion_id):
        queryset = (
            ExpedienteDeclaracion.objects
            .filter(declaracion_id=declaracion_id)
            .select_related('documento__usuario', 'tipo')
        )

        # Agrupamos por ID de tipo (para poder ordenarlo), pero guardamos el nombre también
        grupos_por_tipo = {}

        for expediente in queryset:
            if expediente.tipo:
                tipo_id = expediente.tipo.id
                tipo_nombre = expediente.tipo.nombre
            else:
                tipo_id = 0  # usamos 0 para "Sin clasificar"
                tipo_nombre = "Sin clasificar"

            if tipo_id not in grupos_por_tipo:
                grupos_por_tipo[tipo_id] = {
                    "nombre": tipo_nombre,
                    "documentos": []
                }

            serialized = DocumentoExpedienteSerializer(expediente).data
            grupos_por_tipo[tipo_id]["documentos"].append(serialized)

        # Ordenamos por ID de tipo y reestructuramos como dict con el nombre como clave
        agrupado = OrderedDict()
        for tipo_id in sorted(grupos_por_tipo.keys()):
            nombre = grupos_por_tipo[tipo_id]["nombre"]
            documentos = grupos_por_tipo[tipo_id]["documentos"]
            agrupado[nombre] = documentos

        return Response(agrupado)

class EliminarExpedienteDeclaracionView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            # Obtenemos el expediente
            expediente = ExpedienteDeclaracion.objects.only('id', 'documento_id').get(pk=pk)
            documento_id = expediente.documento_id

            # Eliminamos el expediente primero para romper la relación
            expediente.delete()

            # Obtenemos la ruta del archivo directamente sin cargar el modelo completo
            archivo_path = Documento.objects.filter(pk=documento_id).values_list('archivo', flat=True).first()
            if archivo_path and os.path.isfile(archivo_path):
                os.remove(archivo_path)

            # Eliminamos el documento directamente por ID (sin instanciar el objeto completo)
            Documento.objects.filter(pk=documento_id).delete()

            return Response({'detail': 'Documento y expediente eliminados'}, status=status.HTTP_204_NO_CONTENT)

        except ExpedienteDeclaracion.DoesNotExist:
            return Response({'detail': 'No encontrado'}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return Response({'detail': f'Error al eliminar: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ActualizarEmpresaExpedienteView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, declaracion_id):
        empresa = request.data.get('empresa')

        if not empresa:
            return Response(
                {"detail": "Debe proporcionar empresa"},
                status=status.HTTP_400_BAD_REQUEST
            )

        expedientes = ExpedienteDeclaracion.objects.filter(declaracion_id=declaracion_id)

        if not expedientes.exists():
            return Response(
                {"detail": "No se encontraron expedientes para la declaración especificada"},
                status=status.HTTP_404_NOT_FOUND
            )

        expedientes.update(empresa=empresa)

        return Response(
            {"detail": "empresa actualizada correctamente"},
            status=status.HTTP_200_OK
        )

class ActualizarFolioExpedienteView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        try:
            expediente = ExpedienteDeclaracion.objects.get(pk=pk)
        except ExpedienteDeclaracion.DoesNotExist:
            return Response({'detail': 'No encontrado'}, status=status.HTTP_404_NOT_FOUND)

        serializer = ExpedienteDeclaracionFolioSerializer(expediente, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ActualizarMesAnioFiscalView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, declaracion_id):
        anio_fiscal = request.data.get('anio_fiscal')
        mes_fiscal = request.data.get('mes_fiscal')

        if not anio_fiscal or not mes_fiscal:
            return Response(
                {"detail": "Debe proporcionar anio_fiscal y mes_fiscal"},
                status=status.HTTP_400_BAD_REQUEST
            )

        expedientes = ExpedienteDeclaracion.objects.filter(declaracion_id=declaracion_id)

        if not expedientes.exists():
            return Response(
                {"detail": "No se encontraron expedientes para la declaración especificada"},
                status=status.HTTP_404_NOT_FOUND
            )

        expedientes.update(anio_fiscal=anio_fiscal, mes_fiscal=mes_fiscal)

        return Response(
            {"detail": "Mes y año fiscal actualizados correctamente"},
            status=status.HTTP_200_OK
        )

class TipoDocumentoViewSet(viewsets.ModelViewSet):
    """
    API endpoint para gestionar tipos de documentos.
    Requiere permisos de administrador de importaciones.
    """
    queryset = TipoDocumento.objects.all()
    serializer_class = TipoDocumentoSerializer
    permission_classes = [IsAuthenticated, IsImportacionesAdmin]

class DescargarDocumentosUnificadosPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, declaracion_id):
        expedientes = (
            ExpedienteDeclaracion.objects
            .filter(declaracion_id=declaracion_id)
            .select_related('documento', 'tipo')
        )

        # Agrupar por tipo_id para mantener el orden deseado
        grupos_por_tipo = {}
        for exp in expedientes:
            tipo_id = exp.tipo.id if exp.tipo else 0
            grupos_por_tipo.setdefault(tipo_id, []).append(exp)

        writer = PdfWriter()

        for tipo_id in sorted(grupos_por_tipo.keys()):
            for exp in grupos_por_tipo[tipo_id]:
                doc = exp.documento
                if not doc or not doc.archivo:
                    continue

                try:
                    reader = PdfReader(doc.archivo.path)
                except Exception as e:
                    print(f"Error al leer PDF: {e}")
                    continue

                for page_num in range(len(reader.pages)):
                    page = reader.pages[page_num]

                    # Obtener tamaño real de la página
                    media_box = page.mediabox
                    page_width = float(media_box.width)
                    page_height = float(media_box.height)

                    # Crear overlay con el folio en posición precisa
                    overlay_stream = io.BytesIO()
                    c = canvas.Canvas(overlay_stream, pagesize=(page_width, page_height))
                    c.setFont("Helvetica-Bold", 10)
                    c.setFillColorRGB(1, 0, 0)

                    # Posición del texto: centro horizontal y 0.2 cm desde el borde superior
                    folio_y = page_height - 0.6 * cm
                    c.drawCentredString((page_width / 2)+100, folio_y, exp.folio or "")
                    c.drawCentredString((page_width / 2) - 200, folio_y, f'OC: {exp.orden_compra}  / NI: {exp.nota_ingreso}' or "")
                    c.save()

                    overlay_stream.seek(0)
                    overlay_pdf = PdfReader(overlay_stream)
                    overlay_page = overlay_pdf.pages[0]

                    page.merge_page(overlay_page)
                    writer.add_page(page)

        # Preparar respuesta HTTP
        output_stream = io.BytesIO()
        writer.write(output_stream)
        output_stream.seek(0)

        response = HttpResponse(output_stream.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="documentos_unificados_{declaracion_id}.pdf"'
        return response

class ActualizarOrdenCompraNotaIngresoView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, declaracion_id):
        orden_compra = request.data.get('orden_compra')
        nota_ingreso = request.data.get('nota_ingreso')

        if not orden_compra or not nota_ingreso:
            return Response(
                {"detail": "Debe proporcionar orden de compra y nota de ingreso"},
                status=status.HTTP_400_BAD_REQUEST
            )

        expedientes = ExpedienteDeclaracion.objects.filter(declaracion_id=declaracion_id)

        if not expedientes.exists():
            return Response(
                {"detail": "No se encontraron expedientes para la declaración especificada"},
                status=status.HTTP_404_NOT_FOUND
            )

        expedientes.update(nota_ingreso=nota_ingreso, orden_compra=orden_compra)

        return Response(
            {"detail": "Orden de compra y Nota de ingreso, actualizados correctamente"},
            status=status.HTTP_200_OK
        )

class ActualizarFechaLlegadaFleteView(APIView):
    def post(self, request, pk):
        fecha_str = request.data.get('fecha_llegada')

        if not fecha_str:
            return Response({"error": "Se requiere el campo 'fecha_llegada'."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d")
        except ValueError:
            return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        despacho = get_object_or_404(Despacho, pk=pk)
        despacho.fecha_llegada = fecha
        despacho.save()

        return Response({"message": "Fecha de llegada actualizada correctamente."}, status=status.HTTP_200_OK)


HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/91.0.4472.124 Safari/537.36',
    'Referer': 'https://servicios.senasa.gob.pe/ConsultaRecibos/'
}


class SenasaConsultaView(APIView):
    def post(self, request):
        expediente = request.data.get('expediente')
        ruc = request.data.get('ruc', '')

        print(f"--- INICIO CONSULTA SENASA: {expediente} ---")

        if not expediente:
            return Response({"error": "Falta expediente"}, status=400)

        URL_SEARCH = "https://servicios.senasa.gob.pe/ConsultaRecibos/consulta"

        # Headers para parecer un navegador real
        HEADERS = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36',
            'Referer': 'https://servicios.senasa.gob.pe/ConsultaRecibos/',
            'Origin': 'https://servicios.senasa.gob.pe',
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'
        }

        try:
            payload = {
                'C': 'RECIBO', 'S': 'SEARCH',
                'pcodigoexpediente': expediente, 'pruc': ruc,
                'start': 0, 'limit': 50
            }

            # 1. AGREGAMOS TIMEOUT: Si SENASA no responde en 15 seg, cortamos.
            print("Enviando petición a SENASA...")
            resp = requests.post(URL_SEARCH, data=payload, headers=HEADERS, timeout=15)

            print(f"Respuesta SENASA Status: {resp.status_code}")

            # Si el status no es 200, lanzamos error manual
            resp.raise_for_status()

            raw_data = resp.json()

            # Validación de respuesta vacía o error
            if isinstance(raw_data, dict) and not raw_data:
                print("SENASA devolvió diccionario vacío")
                return Response({"msg": "No se encontraron datos", "detalles": []})

            resultados = []

            # Procesamiento de la lista
            if isinstance(raw_data, list):
                for row in raw_data:
                    # Aseguramos que existan los índices para no crashear
                    if len(row) > 4:
                        nro_recibo = row[0]
                        monto = row[1]
                        fecha = row[3]
                        ucmid_sucio = row[4]

                        ucmid = ucmid_sucio.replace('|', '') if ucmid_sucio else None

                        if ucmid:
                            resultados.append({
                                "nro": nro_recibo,
                                "monto": monto,
                                "fecha": fecha,
                                "ucmid": ucmid,
                                "expediente": expediente,  # Devolvemos esto para usarlo en la descarga
                                "ruc": ruc,  # Devolvemos esto para usarlo en la descarga
                                "status": "Disponible"
                            })

            print(f"Resultados procesados: {len(resultados)}")
            return Response({"msg": "Búsqueda exitosa", "detalles": resultados})

        except requests.exceptions.Timeout:
            print("ERROR: Timeout - SENASA tardó demasiado en responder.")
            return Response({"error": "El servidor de SENASA está lento o caído. Intente de nuevo."}, status=504)

        except requests.exceptions.ConnectionError:
            print("ERROR: ConnectionError - No se pudo conectar a SENASA.")
            return Response({"error": "No hay conexión con SENASA (Posible bloqueo de IP)."}, status=502)

        except Exception as e:
            # ESTO ES LO QUE NECESITAMOS VER EN LA CONSOLA
            print("\n" + "=" * 30)
            print("CRASH EN BACKEND - TRAZA COMPLETA:")
            print(traceback.format_exc())
            print("=" * 30 + "\n")
            return Response({"error": f"Error interno: {str(e)}"}, status=500)


class SenasaDescargaProxyView(APIView):
    def get(self, request):
        # Ahora pedimos el Expediente y el Nro Recibo para regenerar la sesión
        expediente = request.query_params.get('expediente')
        nro_recibo = request.query_params.get('nro')
        ruc = request.query_params.get('ruc', '')

        if not expediente or not nro_recibo:
            return Response({"error": "Faltan datos (expediente o nro recibo)"}, status=400)

        URL_SEARCH = "https://servicios.senasa.gob.pe/ConsultaRecibos/consulta"
        URL_DOWNLOAD = "https://servicios.senasa.gob.pe/sig/upload"

        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/91.0.4472.124 Safari/537.36',
            'Referer': 'https://servicios.senasa.gob.pe/ConsultaRecibos/',
            # Headers importantes para que Java no bloquee la descarga
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        })

        try:
            # PASO 1: RE-AUTENTICACIÓN (Hacer la búsqueda para obtener contexto y ID válido)
            #print(f"--- [Proxy] Re-consultando expediente {expediente} para validar sesión ---")
            payload = {
                'C': 'RECIBO', 'S': 'SEARCH',
                'pcodigoexpediente': expediente, 'pruc': ruc,
                'start': 0, 'limit': 50
            }

            search_resp = session.post(URL_SEARCH, data=payload)
            raw_data = search_resp.json()

            # Buscamos el ID correcto (ucmid) para este recibo específico
            target_ucmid = None

            if isinstance(raw_data, list):
                for row in raw_data:
                    # row[0] es el número de recibo, row[4] es el UCMID
                    if str(row[0]) == str(nro_recibo):
                        target_ucmid = row[4]  # TOMAMOS EL ID CRUDO (CON EL PIPE |)
                        break

            if not target_ucmid:
                return Response({"error": "No se encontró el recibo o ID en SENASA"}, status=404)

            # PASO 2: DESCARGAR CON EL ID EXACTO
            #print(f"--- [Proxy] Descargando ID: {target_ucmid} ---")

            params = {
                'C': 'DLWUCM',
                'fn': random.random(),
                'idx': target_ucmid,  # Enviamos "|5191..." tal cual
                'fns': f"{nro_recibo}.pdf"
            }

            # Usamos la MISMA session que hizo la búsqueda
            pdf_resp = session.get(URL_DOWNLOAD, params=params, stream=True)

            if pdf_resp.status_code == 200:
                # Verificación final de que es un PDF y no un error Java disfrazado
                chunk_inicial = next(pdf_resp.iter_content(64))  # Leemos los primeros bytes

                # Si empieza con %PDF es éxito
                if b'%PDF' in chunk_inicial:
                    # Reconstruimos el generador para la respuesta completa
                    def file_iterator():
                        yield chunk_inicial
                        yield from pdf_resp.iter_content(chunk_size=8192)

                    response = HttpResponse(file_iterator(), content_type='application/pdf')
                    response['Content-Disposition'] = f'attachment; filename="{nro_recibo}.pdf"'
                    return response
                else:
                    # Si no es PDF, devolvemos el error que mandó SENASA para debug
                    return Response({
                        "error": "SENASA devolvió contenido inválido",
                        "content": chunk_inicial.decode('utf-8', errors='ignore')
                    }, status=502)

            return Response({"error": "Error HTTP en descarga"}, status=pdf_resp.status_code)

        except Exception as e:
            print(f"ERROR: {e}")
            return Response({"error": str(e)}, status=500)









