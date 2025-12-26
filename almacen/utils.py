# utils.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from django.utils import timezone


def generate_kardex_excel(data, f_inicio, f_fin, nombre_empresa=""):
    # 1. Crear el libro y la hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Kardex"

    # --- ESTILOS ---
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="005f73", end_color="005f73", fill_type="solid")

    prod_title_font = Font(bold=True, size=11)
    prod_title_fill = PatternFill(start_color="e9ecef", end_color="e9ecef", fill_type="solid")
    prod_border_left = Border(left=Side(style='thick', color='0a9396'))

    table_head_font = Font(bold=True, color="000000", size=10)
    table_head_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    number_format = '#,##0.00'
    date_format = 'dd/mm/yyyy'

    # 2. Encabezado General
    ws.merge_cells('A1:F1')
    cell_title = ws['A1']
    # Concatenamos el nombre de la empresa
    cell_title.value = f"{nombre_empresa} - REPORTE DE KARDEX ({f_inicio} al {f_fin})"
    cell_title.font = header_font
    cell_title.fill = header_fill
    cell_title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    row_num = 3

    # 3. Iterar sobre los productos
    for p_id, producto in data.items():

        # --- A. Título del Producto ---
        product_text = f"{producto['codigo_producto']} - {producto['nombre_producto']}"
        unidad_text = f"Unidad: {producto.get('unidad_medida', 'KG')}"

        ws.merge_cells(f'A{row_num}:F{row_num}')
        cell_prod = ws[f'A{row_num}']
        cell_prod.value = f"{product_text}   |   {unidad_text}"
        cell_prod.font = prod_title_font
        cell_prod.fill = prod_title_fill
        cell_prod.border = prod_border_left

        row_num += 1

        # --- B. Cabeceras de la Tabla ---
        headers = ["Fecha", "Documento", "Detalle", "Entrada", "Salida", "Saldo"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            cell.font = table_head_font
            cell.fill = table_head_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        row_num += 1

        # --- C. Filas de Movimientos ---
        if not producto['kardex']:
            ws.merge_cells(f'A{row_num}:F{row_num}')
            ws[f'A{row_num}'].value = "Sin movimientos en este periodo"
            ws[f'A{row_num}'].font = Font(italic=True, color="999999")
            ws[f'A{row_num}'].alignment = Alignment(horizontal='center')
            row_num += 1
        else:
            for mov in producto['kardex']:

                # --- CORRECCIÓN DE FECHA AQUÍ ---
                fecha_val = mov['fecha']
                # Si es una fecha con zona horaria (UTC), la convertimos a local y quitamos la info de zona
                if hasattr(fecha_val, 'tzinfo') and fecha_val.tzinfo:
                    # 1. Convertir a hora local (Perú)
                    fecha_val = timezone.localtime(fecha_val)
                    # 2. Quitar la metadata de timezone para que Excel no se queje
                    fecha_val = fecha_val.replace(tzinfo=None)

                # Escribir Fecha corregida
                ws.cell(row=row_num, column=1, value=fecha_val).number_format = date_format

                # Resto de columnas (igual que antes)
                ws.cell(row=row_num, column=2, value=mov.get('doc', '-'))
                ws.cell(row=row_num, column=3, value=mov['detalle'])

                c_ent = ws.cell(row=row_num, column=4, value=mov['entrada'] if mov['entrada'] > 0 else "")
                c_ent.number_format = number_format
                if mov['entrada'] > 0: c_ent.font = Font(color="2a9d8f", bold=True)

                c_sal = ws.cell(row=row_num, column=5, value=mov['salida'] if mov['salida'] > 0 else "")
                c_sal.number_format = number_format
                if mov['salida'] > 0: c_sal.font = Font(color="e76f51", bold=True)

                c_bal = ws.cell(row=row_num, column=6, value=mov['saldo'])
                c_bal.number_format = number_format
                c_bal.font = Font(bold=True)

                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).border = thin_border

                row_num += 1

        row_num += 2

    # 4. Ajuste automático
    column_widths = [12, 18, 40, 12, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 5. Guardar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Kardex_{f_inicio}_{f_fin}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_kardex_pdf(data, context, nombre_empresa=""):
    """
    Recibe la data y genera PDF usando un template HTML
    """
    # Renderizamos el HTML enviando 'kardex_items' al template
    html_string = render_to_string('reportes/kardex_pdf.html', {
        'kardex_items': data,
        **context
    })

    output = BytesIO()
    HTML(string=html_string).write_pdf(target=output)
    output.seek(0)

    filename = f"Kardex_{context['fecha_inicio']}.pdf"
    response = HttpResponse(output, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response